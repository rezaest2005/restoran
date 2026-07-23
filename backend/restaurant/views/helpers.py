"""
Shared helper functions and constants.
"""
from __future__ import annotations

import csv
import json as json_module
import logging
import re
from decimal import Decimal
from io import StringIO

import openpyxl
from django.db.models import F
from django.utils import timezone

from ..models import (
    Category, Food, KitchenProduct, RawMaterial,
    ReadyMaterial, SemiFinished, SemiFinishedIngredient,
    PurchaseInvoice, PurchaseInvoiceItem, Supplier,
    InventoryMovement, InventoryUsageLog,
)

logger = logging.getLogger(__name__)


# ═══════════════════════════════════════════════════════════════════
#  CONSTANTS
# ═══════════════════════════════════════════════════════════════════

_DRINK_KEYWORDS = [
    "نوشابه", "دلستر", "دوغ", "ماءالشعیر", "ماء الشعیر", "آب معدنی",
]

_SUPPORTED_ENCODINGS = ("utf-8-sig", "utf-8", "cp1256", "latin-1")

_UNIT_MAP = {
    "کیلوگرم": "kg", "کیلو": "kg", "kg": "kg",
    "گرم": "g", "g": "g",
    "لیتر": "l", "l": "l",
    "میلی‌لیتر": "ml", "میلی لیتر": "ml", "ml": "ml",
    "عدد": "unit", "دسته": "bunch", "بسته": "pack",
}

_UNIT_MAP_FA = {
    "کیلوگرم": "kg", "کیلو": "kg", "گرم": "g",
    "لیتر": "l", "میلی‌لیتر": "ml", "میلی لیتر": "ml",
    "عدد": "unit", "دسته": "bunch", "بسته": "pack",
}

_COL_KEYWORDS = {
    "item_name": ["نام کالا", "نام", "کالا", "ماده اولیه", "item", "name"],
    "quantity": ["مقدار", "تعداد", "quantity", "qty"],
    "unit": ["واحد", "unit"],
    "unit_price": ["قیمت واحد", "قیمت", "فی", "unit_price", "price"],
    "supplier": [
        "تأمین‌کننده", "تامین‌کننده", "تامین کننده", "تأمین کننده",
        "شرکت", "supplier", "فروشنده", "توزیع کننده", "توزیع‌کننده",
        "پخش", "نام شرکت",
    ],
}

VALID_WASTE_REASONS = [
    'expired', 'damaged', 'overcooked',
    'quality_issue', 'returned', 'other',
]


# ═══════════════════════════════════════════════════════════════════
#  FOOD & DISCOUNT HELPERS
# ═══════════════════════════════════════════════════════════════════

def _is_drink(name: str) -> bool:
    return any(k in name for k in _DRINK_KEYWORDS)


def _get_food_discount_info(kitchen_product) -> dict | None:
    now = timezone.now()
    for disc in kitchen_product.discounts.filter(is_active=True):
        if disc.expires_at and disc.expires_at <= now:
            disc.is_active = False
            disc.save(update_fields=["is_active"])

    for disc in kitchen_product.discounts.filter(is_active=True):
        if disc.scope == "happy_hour" and disc.start_time and disc.end_time:
            if not (disc.start_time <= now.time() <= disc.end_time):
                continue

        inv = getattr(kitchen_product, "inventory_record", None)
        current_stock = inv.available_quantity if inv else 0
        kitchen_price = int(kitchen_product.selling_price)

        if disc.scope == "inventory_based" and disc.minimum_stock:
            if current_stock >= disc.minimum_stock:
                continue

        discounted = disc.get_discounted_price(kitchen_price, 1, current_stock)
        if discounted < kitchen_price:
            return {
                "name": disc.name,
                "discount_type": disc.discount_type,
                "value": int(disc.value),
                "discounted_price": int(discounted),
                "max_quantity": disc.max_quantity or None,
                "scope": disc.scope,
            }
    return None


def _build_food_entry(food: Food) -> dict:
    kitchen_product = None
    kitchen_price = 0
    discount_info = None

    if hasattr(food, "recipe") and food.recipe:
        kp = food.recipe.kitchen_products.first()
        if kp:
            kitchen_product = kp
            kitchen_price = int(kp.selling_price)
            discount_info = _get_food_discount_info(kp)

    if not kitchen_product:
        kp = KitchenProduct.objects.filter(name=food.name).first()
        if kp:
            kitchen_product = kp
            kitchen_price = int(kp.selling_price)
            discount_info = _get_food_discount_info(kp)

    return {
        "id": food.id,
        "name": food.name,
        "category_id": food.category_id,
        "category_name": food.category.name,
        "final_price": int(food.final_price),
        "kitchen_price": kitchen_price,
        "has_kitchen": kitchen_product is not None,
        "discount": discount_info,
        "image": food.image.url if food.image else "",
    }


def _build_foods_with_discounts() -> tuple[list[dict], list[dict]]:
    foods = Food.objects.select_related("category").all().order_by(
        "category__order", "name"
    )
    categories = Category.objects.filter(is_active=True).order_by("order")

    foods_data = [_build_food_entry(f) for f in foods]
    categories_data = [{"id": c.id, "name": c.name} for c in categories]
    existing_names = {c["name"]: c["id"] for c in categories_data}

    ready_materials = ReadyMaterial.objects.filter(
        quantity__gt=0
    ).exclude(category__isnull=True).select_related("category")

    for rm in ready_materials:
        foods_data.append({
            "id": f"ready_{rm.id}",
            "name": rm.name,
            "category_id": rm.category_id,
            "category_name": rm.category.name,
            "final_price": int(rm.selling_price or 0),
            "kitchen_price": int(rm.selling_price or 0),
            "has_kitchen": False,
            "discount": None,
            "image": "",
            "is_ready": True,
        })
        if rm.category.name not in existing_names:
            categories_data.append({"id": rm.category_id, "name": rm.category.name})
            existing_names[rm.category.name] = rm.category_id

    return foods_data, categories_data


# ═══════════════════════════════════════════════════════════════════
#  WAREHOUSE HELPERS
# ═══════════════════════════════════════════════════════════════════

def _merge_warehouse_data() -> dict[str, dict]:
    merged: dict[str, dict] = {}

    for material in RawMaterial.objects.all():
        key = material.name.strip().lower()
        merged[key] = {
            "name": material.name,
            "label": material.label,
            "unit": material.get_unit_display(),
            "unit_raw": material.unit,
            "quantity": int(material.quantity),
            "price": int(material.price),
            "total": int(material.total_price),
            "sources": ["انبار"],
        }

    for item in PurchaseInvoiceItem.objects.select_related("invoice").all():
        key = item.item_name.strip().lower()
        qty = int(item.quantity)
        price = int(item.unit_price)
        if key in merged:
            if "فاکتور خرید" not in merged[key]["sources"]:
                merged[key]["sources"].append("فاکتور خرید")
        else:
            merged[key] = {
                "name": item.item_name,
                "label": item.invoice.supplier_name,
                "unit": item.get_unit_display(),
                "unit_raw": item.unit,
                "quantity": qty,
                "price": price,
                "total": qty * price,
                "sources": ["فاکتور خرید"],
            }

    return merged


def _update_raw_material_stock(name: str, qty: float, unit: str, price: int):
    mat = RawMaterial.objects.filter(name__iexact=name).first()
    if mat:
        old_stock = float(mat.quantity)
        mat.quantity = old_stock + float(qty)
        mat.price = price
        mat.save()

        InventoryMovement.objects.create(
            raw_material=mat,
            movement_type='in',
            quantity=qty,
            previous_stock=old_stock,
            new_stock=mat.quantity,
            reference_type='PurchaseInvoice',
            notes='ثبت از فاکتور خرید',
        )
    else:
        RawMaterial.objects.create(
            name=name, label="", price=price, unit=unit, quantity=int(qty)
        )


# ═══════════════════════════════════════════════════════════════════
#  INVOICE HELPERS
# ═══════════════════════════════════════════════════════════════════

def _build_invoice_from_post(request) -> 'PurchaseInvoice':
    supplier_name = request.POST.get("supplier_name", "").strip()
    date = request.POST.get("date")
    if not supplier_name:
        raise ValueError("نام تأمین‌کننده الزامی است.")
    if not date:
        raise ValueError("تاریخ فاکتور الزامی است.")

    invoice = PurchaseInvoice.objects.create(
        supplier_name=supplier_name,
        invoice_number=request.POST.get("invoice_number", "").strip(),
        date=date,
        description=request.POST.get("description", "").strip(),
    )
    uploaded_file = request.FILES.get("invoice_file")
    if uploaded_file:
        invoice.file = uploaded_file
        invoice.save()
    return invoice


def _attach_invoice_items(invoice, post_data) -> int:
    item_names = post_data.getlist("item_name")
    quantities = post_data.getlist("quantity")
    units = post_data.getlist("unit")
    unit_prices = post_data.getlist("unit_price")
    categories = post_data.getlist("category")
    created = 0

    for i, name_raw in enumerate(item_names):
        name = name_raw.strip()
        if not name or name in ("جمع کل", "جمع"):
            continue

        qty = float(quantities[i]) if i < len(quantities) and quantities[i] else 0
        unit = units[i] if i < len(units) else "unit"
        price = (
            int(float(unit_prices[i]))
            if i < len(unit_prices) and unit_prices[i]
            else 0
        )

        category = None
        if i < len(categories) and categories[i]:
            try:
                category = Category.objects.get(id=int(categories[i]))
            except (Category.DoesNotExist, ValueError):
                pass

        if qty > 0 and price > 0:
            PurchaseInvoiceItem.objects.create(
                invoice=invoice,
                item_name=name,
                quantity=qty,
                unit=unit,
                unit_price=price,
                category=category,
            )
            _update_raw_material_stock(name, qty, unit, price)
            created += 1

    return created


# ═══════════════════════════════════════════════════════════════════
#  EXCEL / CSV PARSING
# ═══════════════════════════════════════════════════════════════════

def _read_file_rows(f) -> list[list[str]]:
    filename = f.name.lower()
    if filename.endswith(".csv"):
        return _read_csv_rows(f)
    if filename.endswith((".xlsx", ".xls")):
        return _read_xlsx_rows(f)
    raise ValueError("فرمت فایل پشتیبانی نمی‌شود.")


def _read_csv_rows(f) -> list[list[str]]:
    raw = f.read()
    text = None
    for encoding in _SUPPORTED_ENCODINGS:
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("encoding فایل قابل خواندن نیست.")
    dialect = csv.Sniffer().sniff(text[:1024], delimiters=",;\t")
    reader = csv.reader(StringIO(text), dialect)
    return [row for row in reader if any(cell.strip() for cell in row)]


def _read_xlsx_rows(f) -> list[list[str]]:
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    if not ws or ws.max_row < 2:
        raise ValueError("فایل اکسل خالی است.")
    return [
        [str(cell) if cell is not None else "" for cell in row]
        for row in ws.iter_rows(values_only=True)
    ]


def _detect_column_map(headers: list[str]) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for i, header in enumerate(headers):
        for field, keywords in _COL_KEYWORDS.items():
            if field not in col_map and any(kw in header for kw in keywords):
                col_map[field] = i
    if "item_name" not in col_map:
        col_map = {"item_name": 0, "quantity": 1, "unit": 2, "unit_price": 3}
    return col_map


def _extract_items_from_rows(rows: list[list[str]]) -> tuple[list[dict], str]:
    if len(rows) < 2:
        raise ValueError("فایل خالی است.")
    headers = [str(h).strip().lower() for h in rows[0]]
    col_map = _detect_column_map(headers)
    items = []
    supplier_name = ""

    for row in rows[1:]:
        if not row or all(str(cell).strip() == "" for cell in row):
            continue
        name = _get_cell_str(row, col_map.get("item_name"))
        if not name or name in ("جمع کل", "جمع"):
            continue
        if not supplier_name and "supplier" in col_map:
            sup = _get_cell_str(row, col_map.get("supplier"))
            if sup and "جمع" not in sup:
                supplier_name = sup
        items.append({
            "item_name": name,
            "quantity": _get_cell_float(row, col_map.get("quantity")),
            "unit": _UNIT_MAP.get(
                _get_cell_str(row, col_map.get("unit")), "unit"
            ),
            "unit_price": _get_cell_int(row, col_map.get("unit_price")),
        })
    return items, supplier_name


def _get_cell_str(row: list, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    return str(row[idx]).strip()


def _get_cell_float(row: list, idx: int | None) -> float:
    try:
        return float(_get_cell_str(row, idx))
    except (ValueError, TypeError):
        return 0.0


def _get_cell_int(row: list, idx: int | None) -> int:
    try:
        return int(float(_get_cell_str(row, idx)))
    except (ValueError, TypeError):
        return 0


# ═══════════════════════════════════════════════════════════════════
#  SEMI-FINISHED HELPERS
# ═══════════════════════════════════════════════════════════════════

def _get_or_sync_ingredients(sf) -> list[dict]:
    ingredients = []

    for ing in SemiFinishedIngredient.objects.filter(
        semi_finished=sf
    ).select_related("raw_material"):
        rm = ing.raw_material
        ingredients.append({
            "id": ing.id,
            "raw_material_id": rm.id,
            "raw_material_name": rm.name,
            "quantity": float(ing.quantity),
            "unit": rm.unit,
            "price": int(rm.price),
            "stock": float(rm.quantity),
        })

    if ingredients:
        return ingredients

    desc = sf.description or ""
    if not desc:
        return []

    desc = re.sub(
        r"^[\s]*مواد(\s+مصرفی)?[\s:]*", "", desc, flags=re.IGNORECASE
    ).strip()
    parts = [p.strip() for p in desc.split("|") if p.strip()]

    for part in parts:
        m = re.match(r"(.+?)\s*$$(.+?)$$\s*:\s*([\d.]+)", part)
        if m:
            name = m.group(1).strip()
            unit_fa = m.group(2).strip()
            qty = float(m.group(3))
        else:
            m2 = re.match(r"(.+?)\s*:\s*([\d.]+)\s*(.+)", part)
            if m2:
                name = m2.group(1).strip()
                qty = float(m2.group(2))
                unit_fa = m2.group(3).strip()
            else:
                continue

        unit_code = _UNIT_MAP_FA.get(unit_fa, "unit")

        rm = RawMaterial.objects.filter(name__icontains=name).first()
        if not rm:
            for length in range(len(name), max(0, len(name) - 4), -1):
                rm = RawMaterial.objects.filter(
                    name__icontains=name[:length]
                ).first()
                if rm:
                    break
        if not rm:
            rm = RawMaterial.objects.create(
                name=name, label="", price=0, unit=unit_code, quantity=0
            )

        sfi, _ = SemiFinishedIngredient.objects.get_or_create(
            semi_finished=sf,
            raw_material=rm,
            defaults={"quantity": qty},
        )

        ingredients.append({
            "id": sfi.id,
            "raw_material_id": rm.id,
            "raw_material_name": rm.name,
            "quantity": float(sfi.quantity),
            "unit": rm.unit,
            "price": int(rm.price),
            "stock": float(rm.quantity),
        })

    return ingredients