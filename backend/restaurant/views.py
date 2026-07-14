"""
Restaurant Management System — Views
=====================================
API ViewSets  ·  Pages  ·  Purchase Invoices
Raw Materials  ·  Suppliers  ·  Warehouse Report
Semi-Finished Products  ·  Finished Products  ·  Usage Log
Kitchen Management  ·  Loyalty  ·  Authentication
"""

from __future__ import annotations

# ──────────────────────────────────────────────────────────────────────────────
#  Standard Library
# ──────────────────────────────────────────────────────────────────────────────
import csv
import json as json_module
import logging
import re
import datetime
from django.db.models import Sum,F, Q
from django.utils import timezone
from decimal import Decimal,InvalidOperation
from io import StringIO
import requests
import json


# ──────────────────────────────────────────────────────────────────────────────
#  Third-Party
# ──────────────────────────────────────────────────────────────────────────────
import openpyxl
from rest_framework import generics, permissions, status, viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework_simplejwt.views import TokenObtainPairView, TokenRefreshView
from rest_framework.permissions import IsAuthenticated
from rest_framework.permissions import AllowAny

# ──────────────────────────────────────────────────────────────────────────────
#  Django
# ──────────────────────────────────────────────────────────────────────────────
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth import get_user_model
from django.core.exceptions import ValidationError
from django.db import transaction
from django.http import HttpRequest, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.views.decorators.csrf import csrf_exempt
from datetime import date as date_cls
from django.conf import settings

logger = logging.getLogger(__name__)

# ──────────────────────────────────────────────────────────────────────────────
#  Local — Permissions
# ──────────────────────────────────────────────────────────────────────────────
from .permissions import IsOwnerOrManagerOrKitchenStaff

# ──────────────────────────────────────────────────────────────────────────────
#  Local — Models
# ──────────────────────────────────────────────────────────────────────────────
from .models import (
    Category,
    Coupon,
    CustomerProfile,
    Food,
    InventoryUsageLog,
    KitchenDiscount,
    KitchenInventory,
    KitchenProduct,
    LoyaltyNotification,
    LoyaltyTransaction,
    LoyaltyWallet,
    MembershipLevel,
    Order,
    OrderItem,
    ProductionLog,
    ProductionPlan,
    PurchaseInvoice,
    PurchaseInvoiceItem,
    RawMaterial,
    ReadyMaterial,
    Recipe,
    Referral,
    Reservation,
    Reward,
    RewardRedemption,
    SemiFinished,
    SemiFinishedIngredient,
    Supplier,
    Table,
    WasteLog,
    InventoryMovement,
)

# ──────────────────────────────────────────────────────────────────────────────
#  Local — Serializers
# ──────────────────────────────────────────────────────────────────────────────
from .serializers import (
    CategorySerializer,
    ChangePasswordSerializer,
    CouponApplySerializer,
    CouponCreateSerializer,
    CouponDetailSerializer,
    CouponListSerializer,
    CouponValidateSerializer,
    CustomTokenObtainSerializer,
    CustomerCreateSerializer,
    CustomerDetailSerializer,
    CustomerListSerializer,
    CustomerUpdateSerializer,
    EarnPointsSerializer,
    FoodSerializer,
    KitchenDiscountSerializer,
    KitchenInventorySerializer,
    KitchenProductSerializer,
    LoyaltyDashboardSerializer,
    LoyaltyTransactionSerializer,
    MembershipLevelSerializer,
    NotificationMarkReadSerializer,
    NotificationSerializer,
    OrderSerializer,
    ProcessOrderLoyaltySerializer,
    ProfileSerializer,
    ReadyMaterialSerializer,
    ProductionLogSerializer,
    ProductionPlanSerializer,
    RedeemPointsSerializer,
    ReferralSerializer,
    RegisterSerializer,
    ReservationSerializer,
    ResetPasswordSerializer,
    RewardCreateSerializer,
    RewardDetailSerializer,
    RewardListSerializer,
    RewardRedemptionSerializer,
    SemiFinishedSerializer,
    TableSerializer,
    UserDetailSerializer,
    UserListSerializer,
    WalletDebitSerializer,
    WalletDepositSerializer,
    WalletSerializer,
    WalletTransactionSerializer,
)

# ──────────────────────────────────────────────────────────────────────────────
#  Local — Services
# ──────────────────────────────────────────────────────────────────────────────
from .auth_services import change_password, register_user, reset_password
from .kitchen_services import (
    approve_production_plan,
    calculate_max_production,
    execute_production_plan,
    generate_kitchen_dashboard,
    get_required_materials,
    produce_item,
)
from .services import (
    apply_coupon,
    check_and_grant_birthday_bonus,
    check_level_upgrade,
    earn_points_for_order,
    get_loyalty_dashboard,
    process_order_loyalty,
    redeem_points,
    redeem_reward,
    register_customer,
    run_birthday_check_all,
    seed_membership_levels,
    validate_coupon,
    wallet_deposit,
    wallet_debit,
)
from .utils import api_error, api_success

from restaurant.models import (
    Order, OrderItem, KitchenProduct, KitchenInventory,
    WasteLog, DayCloseReport, DayCloseLog, User,
)
AuthUser = get_user_model()


# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS — Shared utilities
# ══════════════════════════════════════════════════════════════════════════════


def _get_food_discount_info(kitchen_product) -> dict | None:
    """محاسبه اطلاعات تخفیف فعال برای یک KitchenProduct.

    منطق تخفیف قبلاً سه‌بار تکرار شده بود — این تابع جایگزین همه آنهاست.
    """
    now = timezone.now()

    # غیرفعال‌سازی خودکار تخفیف‌های منقضی
    for disc in kitchen_product.discounts.filter(is_active=True):
        if disc.expires_at and disc.expires_at <= now:
            disc.is_active = False
            disc.save(update_fields=["is_active"])

    for disc in kitchen_product.discounts.filter(is_active=True):
        # فیلتر بازه ساعتی
        if disc.scope == "happy_hour" and disc.start_time and disc.end_time:
            if not (disc.start_time <= now.time() <= disc.end_time):
                continue

        inv = getattr(kitchen_product, "inventory_record", None)
        current_stock = inv.available_quantity if inv else 0
        kitchen_price = int(kitchen_product.selling_price)

        # فیلتر بر اساس موجودی انبار
        if disc.scope == "inventory_based" and disc.minimum_stock:
            if current_stock >= disc.minimum_stock:
                continue

        discounted = disc.get_discounted_price(kitchen_price, 1, current_stock)
        if discounted < kitchen_price:
            return {
                "name": disc.name,
                "discount_type": disc.discount_type,
                "value": int(disc.value),
                "discounted_price": int(discounted),
                "max_quantity": disc.max_quantity or None,
                "scope": disc.scope,
            }
    return None


def _build_food_entry(food: Food) -> dict:
    """ساخت دیکشناری یک غذا با اطلاعات آشپزخانه و تخفیف."""
    kitchen_product = None
    kitchen_price = 0
    discount_info = None

    # روش ۱: لینک از طریق recipe
    if hasattr(food, "recipe") and food.recipe:
        kp = food.recipe.kitchen_products.first()
        if kp:
            kitchen_product = kp
            kitchen_price = int(kp.selling_price)
            discount_info = _get_food_discount_info(kp)

    # روش ۲: جستجو با اسم (برای نوشیدنی‌ها و آیتم‌هایی که recipe ندارن)
    if not kitchen_product:
        kp = KitchenProduct.objects.filter(name=food.name).first()
        if kp:
            kitchen_product = kp
            kitchen_price = int(kp.selling_price)
            discount_info = _get_food_discount_info(kp)

    return {
        "id": food.id,
        "name": food.name,
        "category_id": food.category_id,
        "category_name": food.category.name,
        "final_price": int(food.final_price),
        "kitchen_price": kitchen_price,
        "has_kitchen": kitchen_product is not None,
        "discount": discount_info,
        "image": food.image.url if food.image else "",
    }

def _build_foods_with_discounts() -> tuple[list[dict], list[dict]]:
    """ساخت لیست کامل غذاها با تخفیف + مواد آماده دارای دسته‌بندی."""
    foods = Food.objects.select_related("category").all().order_by(
        "category__order", "name"
    )
    categories = Category.objects.filter(is_active=True).order_by("order")

    foods_data = [_build_food_entry(f) for f in foods]
    categories_data = [{"id": c.id, "name": c.name} for c in categories]
    existing_names = {c["name"]: c["id"] for c in categories_data}

    # ── فقط مواد آماده‌ای که دسته‌بندی دارن ──
    ready_materials = ReadyMaterial.objects.filter(
        quantity__gt=0
    ).exclude(category__isnull=True).select_related("category")

    for rm in ready_materials:
        foods_data.append({
            "id": f"ready_{rm.id}",
            "name": rm.name,
            "category_id": rm.category_id,
            "category_name": rm.category.name,
            "final_price": int(rm.selling_price or 0),
            "kitchen_price": int(rm.selling_price or 0),
            "has_kitchen": False,
            "discount": None,
            "image": "",
            "is_ready": True,
        })

        if rm.category.name not in existing_names:
            categories_data.append({"id": rm.category_id, "name": rm.category.name})
            existing_names[rm.category.name] = rm.category_id

    return foods_data, categories_data
# ══════════════════════════════════════════════════════════════════════════════
#  RESTAURANT — API ViewSets
# ══════════════════════════════════════════════════════════════════════════════


class CategoryViewSet(viewsets.ModelViewSet):
    queryset = Category.objects.filter(is_active=True)
    serializer_class = CategorySerializer


class FoodViewSet(viewsets.ModelViewSet):
    queryset = Food.objects.all()
    serializer_class = FoodSerializer

    def get_queryset(self):
        qs = Food.objects.all()
        category = self.request.query_params.get("category")
        if category:
            qs = qs.filter(category=category)
        return qs


class TableViewSet(viewsets.ModelViewSet):
    queryset = Table.objects.all()
    serializer_class = TableSerializer


class ReservationViewSet(viewsets.ModelViewSet):
    queryset = Reservation.objects.all()
    serializer_class = ReservationSerializer


class OrderViewSet(viewsets.ModelViewSet):
    queryset = Order.objects.all().prefetch_related('items__food')
    serializer_class = OrderSerializer
    permission_classes = []  # AllowAny
    authentication_classes = []

    def create(self, request):
        items_data = request.data.get("items", [])
        if not items_data:
            return Response(
                {"error": "آیتمی ارسال نشد."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        for field in ("customer_name", "phone"):
            if field not in request.data or not str(request.data[field]).strip():
                return Response(
                    {"error": f"فیلد «{field}» الزامی است."},
                    status=status.HTTP_400_BAD_REQUEST,
                )

        with transaction.atomic():
            total_price = sum(
                item["price"] * item["quantity"] for item in items_data
            )
            order = Order.objects.create(
                customer_name=request.data["customer_name"],
                phone=request.data["phone"],
                table_id=request.data.get("table"),
                total_price=total_price,
            )
            for item in items_data:
                OrderItem.objects.create(
                    order=order,
                    food_id=item["food"],
                    quantity=item["quantity"],
                    price=item["price"],
                )

        return Response(
            OrderSerializer(order).data,
            status=status.HTTP_201_CREATED,
        )

    def partial_update(self, request, *args, **kwargs):
        order = self.get_object()
        new_status = request.data.get("status")
        valid = ["pending", "confirmed", "preparing", "ready", "delivered", "cancelled"]
        if new_status and new_status in valid:
            order.status = new_status
            order.save()
            return Response(OrderSerializer(order).data)
        return Response(
            {"error": "وضعیت نامعتبر"},
            status=status.HTTP_400_BAD_REQUEST,
        )
# ═══════════════════════════════════════════
#  Dashboard Template View
# ═══════════════════════════════════════════

@staff_member_required
def orders_dashboard(request):
    orders = Order.objects.prefetch_related('items__food').order_by('-created_at')

    # status counts
    pending = 0
    preparing = 0
    ready = 0
    for o in orders:
        if o.status == 'pending':
            pending += 1
        elif o.status == 'preparing':
            preparing += 1
        elif o.status == 'ready':
            ready += 1

    return render(request, 'restaurant/orders_dashboard.html', {
        'orders': orders,
        'pending_count': pending,
        'preparing_count': preparing,
        'ready_count': ready,
    })

@api_view(["POST"])
@staff_member_required
def order_change_status(request, pk):
    try:
        order = Order.objects.get(pk=pk)
    except Order.DoesNotExist:
        return JsonResponse({"error": "سفارش یافت نشد."}, status=404)

    new_status = request.data.get("status")
    valid = ['pending', 'preparing', 'ready', 'delivered', 'cancelled']
    if new_status not in valid:
        return JsonResponse({"error": "وضعیت نامعتبر."}, status=400)

    order.status = new_status
    order.save(update_fields=['status'])
    return JsonResponse({"success": True, "id": order.id, "status": new_status})

class SemiFinishedViewSet(viewsets.ModelViewSet):
    queryset = SemiFinished.objects.all()
    serializer_class = SemiFinishedSerializer


class ReadyMaterialViewSet(viewsets.ModelViewSet):
    queryset = ReadyMaterial.objects.all()
    serializer_class = ReadyMaterialSerializer

    def get_queryset(self):
        qs = ReadyMaterial.objects.select_related("supplier").all()
        search = self.request.query_params.get("search")
        if search:
            qs = qs.filter(
                Q(name__icontains=search) | Q(barcode__icontains=search)
            )
        return qs


# ══════════════════════════════════════════════════════════════════════════════
#  Pages — General
# ══════════════════════════════════════════════════════════════════════════════


@login_required
def home(request: HttpRequest):
    return render(request, "admin/index.html")


def auth_page(request: HttpRequest):
    if request.user.is_authenticated:
        return redirect("home")
    return render(request, "auth.html")


@login_required
def logout_page(request: HttpRequest):
    from django.contrib.auth import logout
    logout(request)
    return redirect("auth_page")


# ══════════════════════════════════════════════════════════════════════════════
#  Purchase Invoices
# ══════════════════════════════════════════════════════════════════════════════


@staff_member_required
def purchase_invoice_list(request: HttpRequest):
    invoices = PurchaseInvoice.objects.all().order_by("-date")
    return render(request, "restaurant/invoice_list.html", {"invoices": invoices})


@staff_member_required
def purchase_invoice_detail(request: HttpRequest, pk: int):
    invoice = get_object_or_404(PurchaseInvoice, pk=pk)
    return render(request, "restaurant/invoice_detail.html", {"invoice": invoice})


@csrf_protect
@staff_member_required
def create_purchase_invoice(request: HttpRequest):
    if request.method == "POST":
        try:
            invoice = _build_invoice_from_post(request)
            created = _attach_invoice_items(invoice, request.POST)
            messages.success(request, f"فاکتور خرید با {created} قلم کالا ثبت شد.")
            return redirect("/admin/restaurant/purchaseinvoice/")
        except ValueError as exc:
            messages.error(request, str(exc))
        except Exception as exc:
            logger.exception("Error creating purchase invoice")
            messages.error(request, f"خطا در ثبت فاکتور: {exc}")
    categories = Category.objects.filter(is_active=True).order_by('order')
    categories_json = json_module.dumps(
        [{'id': c.id, 'name': c.name} for c in categories],
        ensure_ascii=False
    )
    return render(
        request,
        "restaurant/create_invoice.html",
        {
            "unit_choices": RawMaterial.UNIT_CHOICES,
            "categories_json": categories_json,
        },
    )



def _build_invoice_from_post(request: HttpRequest) -> PurchaseInvoice:
    """ساخت PurchaseInvoice از داده‌های POST."""
    supplier_name = request.POST.get("supplier_name", "").strip()
    date = request.POST.get("date")
    if not supplier_name:
        raise ValueError("نام تأمین‌کننده الزامی است.")
    if not date:
        raise ValueError("تاریخ فاکتور الزامی است.")

    invoice = PurchaseInvoice.objects.create(
        supplier_name=supplier_name,
        invoice_number=request.POST.get("invoice_number", "").strip(),
        date=date,
        description=request.POST.get("description", "").strip(),
    )
    uploaded_file = request.FILES.get("invoice_file")
    if uploaded_file:
        invoice.file = uploaded_file
        invoice.save()
    return invoice


def _attach_invoice_items(invoice: PurchaseInvoice, post_data) -> int:
    """اتصال اقلام فاکتور و آپدیت انبار مواد اولیه."""
    item_names = post_data.getlist("item_name")
    quantities = post_data.getlist("quantity")
    units = post_data.getlist("unit")
    unit_prices = post_data.getlist("unit_price")
    categories = post_data.getlist("category")
    created = 0

    for i, name_raw in enumerate(item_names):
        name = name_raw.strip()
        if not name or name in ("جمع کل", "جمع"):
            continue

        qty = float(quantities[i]) if i < len(quantities) and quantities[i] else 0
        unit = units[i] if i < len(units) else "unit"
        price = (
            int(float(unit_prices[i]))
            if i < len(unit_prices) and unit_prices[i]
            else 0
        )

        # ── خواندن دسته‌بندی ──
        category = None
        if i < len(categories) and categories[i]:
            try:
                category = Category.objects.get(id=int(categories[i]))
            except (Category.DoesNotExist, ValueError):
                pass

        if qty > 0 and price > 0:
            PurchaseInvoiceItem.objects.create(
                invoice=invoice,
                item_name=name,
                quantity=qty,
                unit=unit,
                unit_price=price,
                category=category,
            )
            _update_raw_material_stock(name, qty, unit, price)
            created += 1

    return created


def _update_raw_material_stock(name: str, qty: float, unit: str, price: int):
    """اگر ماده اولیه وجود دارد → موجودی اضافه + قیمت آپدیت + جابجایی انبار."""
    mat = RawMaterial.objects.filter(name__iexact=name).first()
    if mat:
        old_price = mat.price
        old_stock = float(mat.quantity)

        mat.quantity = old_stock + float(qty)
        mat.price = price
        mat.save()

        InventoryMovement.objects.create(
            raw_material   = mat,
            movement_type  = 'in',
            quantity       = qty,
            previous_stock = old_stock,
            new_stock      = mat.quantity,
            reference_type = 'PurchaseInvoice',
            notes          = 'ثبت از فاکتور خرید',
        )
    else:
        RawMaterial.objects.create(
            name=name, label="", price=price, unit=unit, quantity=int(qty)
        )


@csrf_protect
@require_POST
def parse_excel_file(request: HttpRequest):
    """پارس فایل اکسل/CSV فاکتور خرید."""
    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return JsonResponse({"success": False, "error": "فایلی ارسال نشد."})
    try:
        rows = _read_file_rows(uploaded_file)
        items, supplier_name = _extract_items_from_rows(rows)
        if not items:
            return JsonResponse(
                {"success": False, "error": "هیچ کالایی در فایل یافت نشد."}
            )

        supplier_id = None
        if supplier_name:
            sup = Supplier.objects.filter(name__icontains=supplier_name).first()
            if sup:
                supplier_id = sup.id

        return JsonResponse(
            {
                "success": True,
                "items": items,
                "count": len(items),
                "supplier_name": supplier_name,
                "supplier_id": supplier_id,
            }
        )
    except Exception as exc:
        logger.exception("Error parsing excel file")
        return JsonResponse({"success": False, "error": f"خطا: {exc}"})


# ── Excel/CSV parsing helpers ────────────────────────────────────────────────

_SUPPORTED_ENCODINGS = ("utf-8-sig", "utf-8", "cp1256", "latin-1")

_UNIT_MAP = {
    "کیلوگرم": "kg",
    "کیلو": "kg",
    "kg": "kg",
    "گرم": "g",
    "g": "g",
    "لیتر": "l",
    "l": "l",
    "میلی‌لیتر": "ml",
    "میلی لیتر": "ml",
    "ml": "ml",
    "عدد": "unit",
    "دسته": "bunch",
    "بسته": "pack",
}

_COL_KEYWORDS = {
    "item_name": ["نام کالا", "نام", "کالا", "ماده اولیه", "item", "name"],
    "quantity": ["مقدار", "تعداد", "quantity", "qty"],
    "unit": ["واحد", "unit"],
    "unit_price": ["قیمت واحد", "قیمت", "فی", "unit_price", "price"],
    "supplier": [
        "تأمین‌کننده", "تامین‌کننده", "تامین کننده", "تأمین کننده",
        "شرکت", "supplier", "فروشنده", "توزیع کننده", "توزیع‌کننده",
        "پخش", "نام شرکت",
    ],
}


def _read_file_rows(f) -> list[list[str]]:
    filename = f.name.lower()
    if filename.endswith(".csv"):
        return _read_csv_rows(f)
    if filename.endswith((".xlsx", ".xls")):
        return _read_xlsx_rows(f)
    raise ValueError("فرمت فایل پشتیبانی نمی‌شود.")


def _read_csv_rows(f) -> list[list[str]]:
    raw = f.read()
    text = None
    for encoding in _SUPPORTED_ENCODINGS:
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("encoding فایل قابل خواندن نیست.")
    dialect = csv.Sniffer().sniff(text[:1024], delimiters=",;\t")
    reader = csv.reader(StringIO(text), dialect)
    return [row for row in reader if any(cell.strip() for cell in row)]


def _read_xlsx_rows(f) -> list[list[str]]:
    wb = openpyxl.load_workbook(f, data_only=True)
    ws = wb.active
    if not ws or ws.max_row < 2:
        raise ValueError("فایل اکسل خالی است.")
    return [
        [str(cell) if cell is not None else "" for cell in row]
        for row in ws.iter_rows(values_only=True)
    ]


def _detect_column_map(headers: list[str]) -> dict[str, int]:
    col_map: dict[str, int] = {}
    for i, header in enumerate(headers):
        for field, keywords in _COL_KEYWORDS.items():
            if field not in col_map and any(kw in header for kw in keywords):
                col_map[field] = i
    if "item_name" not in col_map:
        col_map = {"item_name": 0, "quantity": 1, "unit": 2, "unit_price": 3}
    return col_map


def _extract_items_from_rows(rows: list[list[str]]) -> tuple[list[dict], str]:
    if len(rows) < 2:
        raise ValueError("فایل خالی است.")
    headers = [str(h).strip().lower() for h in rows[0]]
    col_map = _detect_column_map(headers)
    items = []
    supplier_name = ""

    for row in rows[1:]:
        if not row or all(str(cell).strip() == "" for cell in row):
            continue
        name = _get_cell_str(row, col_map.get("item_name"))
        if not name or name in ("جمع کل", "جمع"):
            continue
        if not supplier_name and "supplier" in col_map:
            sup = _get_cell_str(row, col_map.get("supplier"))
            if sup and "جمع" not in sup:
                supplier_name = sup
        items.append(
            {
                "item_name": name,
                "quantity": _get_cell_float(row, col_map.get("quantity")),
                "unit": _UNIT_MAP.get(
                    _get_cell_str(row, col_map.get("unit")), "unit"
                ),
                "unit_price": _get_cell_int(row, col_map.get("unit_price")),
            }
        )
    return items, supplier_name


def _get_cell_str(row: list, idx: int | None) -> str:
    if idx is None or idx >= len(row):
        return ""
    return str(row[idx]).strip()


def _get_cell_float(row: list, idx: int | None) -> float:
    try:
        return float(_get_cell_str(row, idx))
    except (ValueError, TypeError):
        return 0.0


def _get_cell_int(row: list, idx: int | None) -> int:
    try:
        return int(float(_get_cell_str(row, idx)))
    except (ValueError, TypeError):
        return 0


# ══════════════════════════════════════════════════════════════════════════════
#  Raw Materials
# ══════════════════════════════════════════════════════════════════════════════


@csrf_protect
@require_POST
def raw_material_save(request: HttpRequest):
    """ذخیره یا ویرایش ماده اولیه."""
    try:
        pk = request.POST.get("id")
        name = request.POST.get("name", "").strip()
        label = request.POST.get("label", "").strip()
        price = int(float(request.POST.get("price", 0)))
        unit = request.POST.get("unit", "unit")
        quantity = int(float(request.POST.get("quantity", 0)))

        if not name:
            return JsonResponse({"success": False, "error": "نام کالا الزامی است."})
        if price < 0:
            return JsonResponse(
                {"success": False, "error": "قیمت نمی‌تواند منفی باشد."}
            )
        if quantity < 0:
            return JsonResponse(
                {"success": False, "error": "تعداد نمی‌تواند منفی باشد."}
            )

        if pk:
            mat = get_object_or_404(RawMaterial, pk=pk)
            mat.name = name
            mat.label = label
            mat.price = price
            mat.unit = unit
            mat.quantity = quantity
            mat.save()
            msg = "ویرایش شد."
        else:
            mat = RawMaterial.objects.create(
                name=name, label=label, price=price, unit=unit, quantity=quantity
            )
            msg = "اضافه شد."

        return JsonResponse(
            {
                "success": True,
                "msg": msg,
                "item": {
                    "id": mat.pk,
                    "name": mat.name,
                    "label": mat.label,
                    "price": int(mat.price),
                    "unit": mat.unit,
                    "unit_display": mat.get_unit_display(),
                    "quantity": int(mat.quantity),
                    "total": int(mat.total_price),
                },
            }
        )
    except Exception as exc:
        logger.exception("Error saving raw material")
        return JsonResponse({"success": False, "error": str(exc)})


@csrf_protect
@require_POST
def raw_material_delete(request: HttpRequest):
    try:
        mat = get_object_or_404(RawMaterial, pk=request.POST.get("id"))
        name = mat.name
        mat.delete()
        return JsonResponse({"success": True, "msg": f"«{name}» حذف شد."})
    except Exception as exc:
        logger.exception("Error deleting raw material")
        return JsonResponse({"success": False, "error": str(exc)})


def raw_material_suggestions(request: HttpRequest):
    query = request.GET.get("q", "").strip()
    if not query:
        return JsonResponse([], safe=False)
    materials = (
        RawMaterial.objects.filter(name__icontains=query)
        .values("name", "unit", "price")
        .order_by("name")[:10]
    )
    return JsonResponse(
        [
            {"name": m["name"], "unit": m["unit"], "price": int(m["price"])}
            for m in materials
        ],
        safe=False,
    )


# ══════════════════════════════════════════════════════════════════════════════
#  Suppliers
# ══════════════════════════════════════════════════════════════════════════════


@staff_member_required
def supplier_list(request: HttpRequest):
    suppliers = Supplier.objects.all().values(
        "id", "name", "phone", "address", "contact_person"
    )
    return JsonResponse(list(suppliers), safe=False)


def supplier_suggestions(request: HttpRequest):
    query = request.GET.get("q", "").strip()
    if not query:
        return JsonResponse([], safe=False)
    results = [
        {
            "id": s.id,
            "name": s.name,
            "phone": s.phone or "",
            "address": s.address or "",
            "contact_person": s.contact_person or "",
        }
        for s in Supplier.objects.filter(name__icontains=query)[:10]
    ]
    return JsonResponse(results, safe=False)


@csrf_protect
@require_POST
@staff_member_required
def supplier_save(request: HttpRequest):
    try:
        sup_id = request.POST.get("id", "").strip()
        name = request.POST.get("name", "").strip()
        if not name:
            return JsonResponse(
                {"success": False, "error": "نام شرکت الزامی است."}
            )
        if sup_id:
            supplier = get_object_or_404(Supplier, pk=sup_id)
        else:
            supplier = Supplier()
        supplier.name = name
        supplier.phone = request.POST.get("phone", "").strip()
        supplier.address = request.POST.get("address", "").strip()
        supplier.contact_person = request.POST.get("contact_person", "").strip()
        supplier.description = request.POST.get("description", "").strip()
        supplier.save()
        return JsonResponse(
            {
                "id": supplier.pk,
                "name": supplier.name,
                "phone": supplier.phone or "",
                "address": supplier.address or "",
                "contact_person": supplier.contact_person or "",
                "msg": "تأمین‌کننده ذخیره شد.",
            }
        )
    except Exception as exc:
        logger.exception("Error saving supplier")
        return JsonResponse({"success": False, "error": str(exc)})


@csrf_protect
@require_POST
@staff_member_required
def supplier_delete(request: HttpRequest):
    try:
        pk = request.POST.get("id")
        if not pk:
            return JsonResponse(
                {"success": False, "error": "شناسه ارسال نشد."}
            )
        sup = get_object_or_404(Supplier, pk=pk)
        name = sup.name
        sup.delete()
        return JsonResponse({"success": True, "msg": f"«{name}» حذف شد."})
    except Exception as exc:
        logger.exception("Error deleting supplier")
        return JsonResponse({"success": False, "error": str(exc)})


# ══════════════════════════════════════════════════════════════════════════════
#  Warehouse Report
# ══════════════════════════════════════════════════════════════════════════════


def _merge_warehouse_data() -> dict[str, dict]:
    """ادغام اطلاعات انبار مواد اولیه و فاکتورهای خرید."""
    merged: dict[str, dict] = {}

    for material in RawMaterial.objects.all():
        key = material.name.strip().lower()
        merged[key] = {
            "name": material.name,
            "label": material.label,
            "unit": material.get_unit_display(),
            "unit_raw": material.unit,
            "quantity": int(material.quantity),
            "price": int(material.price),
            "total": int(material.total_price),
            "sources": ["انبار"],
        }

    for item in PurchaseInvoiceItem.objects.select_related("invoice").all():
        key = item.item_name.strip().lower()
        qty = int(item.quantity)
        price = int(item.unit_price)
        if key in merged:
            if "فاکتور خرید" not in merged[key]["sources"]:
                merged[key]["sources"].append("فاکتور خرید")
        else:
            merged[key] = {
                "name": item.item_name,
                "label": item.invoice.supplier_name,
                "unit": item.get_unit_display(),
                "unit_raw": item.unit,
                "quantity": qty,
                "price": price,
                "total": qty * price,
                "sources": ["فاکتور خرید"],
            }

    return merged


@staff_member_required
def warehouse_json(request: HttpRequest):
    """لیست مواد اولیه با موجودی — برای بررسی موجودی در فرانت‌اند."""
    materials = list(
        RawMaterial.objects.all().values("id", "name", "quantity", "unit", "price")
    )
    for m in materials:
        m["quantity"] = float(m["quantity"])
        m["price"] = int(m["price"])
    return JsonResponse(materials, safe=False)


# ══════════════════════════════════════════════════════════════════════════════
#  Staff-Only Page Views
# ══════════════════════════════════════════════════════════════════════════════

@staff_member_required
def create_invoice_view(request: HttpRequest):
    categories = Category.objects.filter(is_active=True).order_by('order')
    categories_json = json_module.dumps(
        [{'id': c.id, 'name': c.name} for c in categories],
        ensure_ascii=False
    )
    return render(
        request,
        "restaurant/create_invoice.html",
        {
            "unit_choices": RawMaterial.UNIT_CHOICES,
            "categories_json": categories_json,
        },
    )

@staff_member_required
def raw_materials_view(request: HttpRequest):
    return render(
        request,
        "restaurant/raw_materials.html",
        {
            "materials": RawMaterial.objects.all().order_by("name"),
            "unit_choices": RawMaterial.UNIT_CHOICES,
        },
    )


# ══════════════════════════════════════════════════════════════════════════════
#  Semi-Finished Products
# ══════════════════════════════════════════════════════════════════════════════


@staff_member_required
def semi_finished_view(request: HttpRequest):
    semi_finished_list = SemiFinished.objects.prefetch_related(
        "ingredients__raw_material"
    ).all()
    raw_materials = RawMaterial.objects.all()
    merged = _merge_warehouse_data()
    raw_materials_json = json_module.dumps(
        sorted(merged.values(), key=lambda x: x["name"]),
        ensure_ascii=False,
    )
    return render(
        request,
        "restaurant/semi_finished.html",
        {
            "semi_finished_list": semi_finished_list,
            "raw_materials": raw_materials,
            "raw_materials_json": raw_materials_json,
        },
    )


@csrf_protect
@require_POST
def semi_finished_save(request: HttpRequest):
    """ساخت ماده نیمه‌آماده و کسر مواد اولیه از انبار."""
    try:
        data = json_module.loads(request.body)
        name = data.get("name", "").strip()
        if not name:
            return JsonResponse({"success": False, "error": "نام الزامی است."})

        sf = SemiFinished.objects.create(
            name=name,
            unit=data.get("unit", "kg"),
            category=data.get("category", "other"),
            description=data.get("description", ""),
            quantity_produced=float(data.get("quantity_produced", 1)),
            profit_percentage=float(data.get("profit_percentage", 0)),
        )

        for ing in data.get("ingredients", []):
            raw_id = ing.get("raw_material_id")
            qty = float(ing["quantity"])
            ing_name = ing.get("name", "").strip()

            if not raw_id or raw_id == 0:
                if ing_name:
                    mat = RawMaterial.objects.filter(name__iexact=ing_name).first()
                    if not mat:
                        mat = RawMaterial.objects.create(
                            name=ing_name,
                            label=ing.get("label", ""),
                            price=int(float(ing.get("price", 0))),
                            unit=ing.get("unit", "kg"),
                            quantity=0,
                        )
                    raw_id = mat.id
                else:
                    continue

            SemiFinishedIngredient.objects.create(
                semi_finished=sf, raw_material_id=raw_id, quantity=qty
            )

            try:
                raw_mat = RawMaterial.objects.get(id=raw_id)
                raw_mat.quantity = max(0, raw_mat.quantity - Decimal(str(qty)))
                raw_mat.save()
                InventoryUsageLog.objects.create(
                    raw_material=raw_mat,
                    usage_type="semi_finished",
                    quantity_used=qty,
                    reference=f"ساخت: {name}",
                    note=f"ماده نیمه‌آماده «{name}» — {qty} {raw_mat.get_unit_display()}",
                )
            except RawMaterial.DoesNotExist:
                pass

        return JsonResponse({"success": True, "msg": "ذخیره شد.", "id": sf.pk})
    except Exception as exc:
        logger.exception("Error saving semi-finished product")
        return JsonResponse({"success": False, "error": str(exc)})


@csrf_protect
@require_POST
@staff_member_required
def semi_finished_delete(request: HttpRequest):
    """حذف ماده نیمه‌آماده و بازگردانی موجودی انبار."""
    try:
        data = json_module.loads(request.body)
        sf_id = data.get("id")
        if not sf_id:
            return JsonResponse(
                {"success": False, "error": "شناسه ارسال نشد."}
            )

        sf = SemiFinished.objects.get(id=sf_id)
        sf_name = sf.name

        # بازگردانی موجودی مواد اولیه
        for ing in sf.ingredients.all():
            raw_mat = ing.raw_material
            raw_mat.quantity += ing.quantity
            raw_mat.save()

        InventoryUsageLog.objects.filter(reference=f"ساخت: {sf_name}").delete()
        sf.delete()

        return JsonResponse(
            {
                "success": True,
                "msg": f"«{sf_name}» حذف شد و موجودی انبار بازگردانی شد.",
            }
        )
    except SemiFinished.DoesNotExist:
        return JsonResponse(
            {"success": False, "error": "ماده نیمه‌آماده پیدا نشد."}
        )
    except Exception as exc:
        logger.exception("Error deleting semi-finished product")
        return JsonResponse({"success": False, "error": str(exc)})


# ── Semi-Finished Produce ────────────────────────────────────────────────────

_UNIT_MAP_FA = {
    "کیلوگرم": "kg",
    "کیلو": "kg",
    "گرم": "g",
    "لیتر": "l",
    "میلی‌لیتر": "ml",
    "میلی لیتر": "ml",
    "عدد": "unit",
    "دسته": "bunch",
    "بسته": "pack",
}


def _get_or_sync_ingredients(sf: SemiFinished) -> list[dict]:
    """خواندن مواد اولیه SemiFinished.

    ابتدا از SemiFinishedIngredient می‌خواند. اگر خالی بود،
    از description استخراج می‌کند و رکورد می‌سازد.
    """
    ingredients = []

    # ── منبع ۱: SemiFinishedIngredient ──
    for ing in SemiFinishedIngredient.objects.filter(
        semi_finished=sf
    ).select_related("raw_material"):
        rm = ing.raw_material
        ingredients.append(
            {
                "id": ing.id,
                "raw_material_id": rm.id,
                "raw_material_name": rm.name,
                "quantity": float(ing.quantity),
                "unit": rm.unit,
                "price": int(rm.price),
                "stock": float(rm.quantity),
            }
        )

    if ingredients:
        return ingredients

    # ── منبع ۲: استخراج از description ──
    desc = sf.description or ""
    if not desc:
        return []

    desc = re.sub(
        r"^[\s]*مواد(\s+مصرفی)?[\s:]*", "", desc, flags=re.IGNORECASE
    ).strip()
    parts = [p.strip() for p in desc.split("|") if p.strip()]

    for part in parts:
        # فرمت: "جعفری (کیلوگرم): 1 کیلوگرم"
        m = re.match(r"(.+?)\s*$$(.+?)$$\s*:\s*([\d.]+)", part)
        if m:
            name = m.group(1).strip()
            unit_fa = m.group(2).strip()
            qty = float(m.group(3))
        else:
            # فرمت جایگزین: "جعفری: 1 کیلوگرم"
            m2 = re.match(r"(.+?)\s*:\s*([\d.]+)\s*(.+)", part)
            if m2:
                name = m2.group(1).strip()
                qty = float(m2.group(2))
                unit_fa = m2.group(3).strip()
            else:
                continue

        unit_code = _UNIT_MAP_FA.get(unit_fa, "unit")

        # جستجوی ماده اولیه
        rm = RawMaterial.objects.filter(name__icontains=name).first()
        if not rm:
            for length in range(len(name), max(0, len(name) - 4), -1):
                rm = RawMaterial.objects.filter(
                    name__icontains=name[:length]
                ).first()
                if rm:
                    break
        if not rm:
            rm = RawMaterial.objects.create(
                name=name, label="", price=0, unit=unit_code, quantity=0
            )

        sfi, _ = SemiFinishedIngredient.objects.get_or_create(
            semi_finished=sf,
            raw_material=rm,
            defaults={"quantity": qty},
        )

        ingredients.append(
            {
                "id": sfi.id,
                "raw_material_id": rm.id,
                "raw_material_name": rm.name,
                "quantity": float(sfi.quantity),
                "unit": rm.unit,
                "price": int(rm.price),
                "stock": float(rm.quantity),
            }
        )

    return ingredients


@staff_member_required
def semi_finished_produce(request: HttpRequest):
    """تولید ماده نیمه‌آماده و کسر مواد اولیه از انبار."""
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "POST only"})
    try:
        data = json_module.loads(request.body)
        sf_id = data.get("semi_finished_id")
        quantity = float(data.get("quantity", 0))

        if not sf_id:
            return JsonResponse(
                {"success": False, "error": "شناسه ارسال نشد."}
            )
        if quantity <= 0:
            return JsonResponse(
                {"success": False, "error": "تعداد باید بیشتر از صفر باشد."}
            )

        sf = SemiFinished.objects.prefetch_related(
            "ingredients__raw_material"
        ).get(id=sf_id)

        ingredients = sf.ingredients.all()
        if not ingredients.exists():
            return JsonResponse(
                {
                    "success": False,
                    "error": f"هیچ ماده اولیه‌ای برای «{sf.name}» تعریف نشده.",
                }
            )

        # بررسی موجودی
        shortages = []
        for ing in ingredients:
            needed = ing.quantity * Decimal(str(quantity))
            if ing.raw_material.quantity < needed:
                shortages.append(
                    f"«{ing.raw_material.name}»: نیاز {needed} — "
                    f"موجودی {ing.raw_material.quantity}"
                )

        if shortages:
            return JsonResponse(
                {
                    "success": False,
                    "error": "موجودی کافی نیست:\n" + "\n".join(shortages),
                }
            )

        # کسر از انبار + اضافه کردن موجودی نیمه‌آماده
        with transaction.atomic():
            for ing in ingredients:
                needed = ing.quantity * Decimal(str(quantity))
                raw_mat = ing.raw_material
                raw_mat.quantity -= needed
                raw_mat.save()
                InventoryUsageLog.objects.create(
                    raw_material=raw_mat,
                    usage_type="semi_finished",
                    quantity_used=float(needed),
                    reference=f"تولید: {sf.name} × {quantity}",
                    note=f"تولید {quantity} واحد «{sf.name}»",
                )

            # ★ اضافه کردن موجودی نیمه‌آماده
            produced_amount = float(sf.quantity_produced) * quantity
            sf.current_stock += Decimal(str(produced_amount))
            sf.save(update_fields=['current_stock'])

        return JsonResponse(
            {
                "success": True,
                "msg": (
                    f"{int(quantity)} واحد «{sf.name}» تولید شد. "
                    f"({produced_amount} {sf.get_unit_display()} به موجودی اضافه شد. "
                    f"موجودی فعلی: {sf.current_stock})"
                ),
            }
        )
    except SemiFinished.DoesNotExist:
        return JsonResponse(
            {"success": False, "error": "ماده نیمه‌آماده پیدا نشد."}
        )
    except Exception as exc:
        logger.exception("Error producing semi-finished")
        return JsonResponse({"success": False, "error": str(exc)})

@staff_member_required
def semi_finished_produce_detail(request: HttpRequest, pk: int):
    """جزئیات نیمه‌آماده برای صفحه تولید — با استخراج خودکار مواد."""
    sf = get_object_or_404(SemiFinished, pk=pk)
    ingredients = _get_or_sync_ingredients(sf)
    return JsonResponse(
        {
            "id": sf.id,
            "name": sf.name,
            "category": sf.category,
            "description": sf.description or "",
            "unit": sf.unit,
            "quantity_produced": float(sf.quantity_produced or 1),
            "profit_percentage": float(sf.profit_percentage or 0),
            "cost_per_unit": int(sf.cost_per_unit or 0),
            "suggested_price": int(sf.suggested_price or 0),
            "ingredients": ingredients,
        }
    )


# ══════════════════════════════════════════════════════════════════════════════
#  Usage Log
# ══════════════════════════════════════════════════════════════════════════════


@staff_member_required
def usage_log_view(request: HttpRequest):
    semi_finished_list = (
        SemiFinished.objects.prefetch_related("ingredients__raw_material")
        .all()
        .order_by("name")
    )
    total_logs = InventoryUsageLog.objects.count()

    by_type = {}
    for choice_val, choice_label in InventoryUsageLog.USAGE_TYPE_CHOICES:
        by_type[choice_val] = {
            "label": choice_label,
            "count": InventoryUsageLog.objects.filter(
                usage_type=choice_val
            ).count(),
        }

    semi_finished_data = []
    for sf in semi_finished_list:
        ingredients = [
            {
                "name": ing.raw_material.name,
                "unit": ing.raw_material.get_unit_display(),
                "quantity": str(ing.quantity),
                "stock": str(ing.raw_material.quantity),
                "price": str(ing.raw_material.price),
                "total_cost": str(int(ing.total_cost)),
            }
            for ing in sf.ingredients.all()
        ]
        semi_finished_data.append(
            {
                "id": sf.id,
                "name": sf.name,
                "category": sf.get_category_display(),
                "unit": sf.get_unit_display(),
                "quantity_produced": str(sf.quantity_produced),
                "total_cost": str(int(sf.total_cost)),
                "cost_per_unit": str(int(sf.cost_per_unit)),
                "suggested_price": str(int(sf.suggested_price)),
                "can_produce": sf.can_produce,
                "ingredients": ingredients,
            }
        )

    suppliers = Supplier.objects.all().order_by("name")
    order_stats = {
        "total_orders": 0,
        "total_revenue": 0,
        "top_foods": [],
    }

    return render(
        request,
        "restaurant/usage_log.html",
        {
            "semi_finished_list": semi_finished_list,
            "semi_finished_json": json_module.dumps(
                semi_finished_data, ensure_ascii=False
            ),
            "total_logs": total_logs,
            "by_type": by_type,
            "type_choices": InventoryUsageLog.USAGE_TYPE_CHOICES,
            "suppliers": suppliers,
            "order_stats": order_stats,
        },
    )


@staff_member_required
def usage_log_json(request: HttpRequest):
    logs = (
        InventoryUsageLog.objects.select_related("raw_material")
        .all()
        .order_by("-used_at")[:200]
    )
    data = [
        {
            "id": log.id,
            "material": log.raw_material.name,
            "unit": log.raw_material.get_unit_display(),
            "quantity": str(log.quantity_used),
            "type": log.get_usage_type_display(),
            "type_key": log.usage_type,
            "reference": log.reference or "—",
            "note": log.note or "",
            "date": log.used_at.strftime("%Y/%m/%d %H:%M"),
        }
        for log in logs
    ]
    return JsonResponse({"logs": data})


@staff_member_required
def usage_log_detail_json(request: HttpRequest):
    material_id = request.GET.get("material_id", "")
    if not material_id:
        return JsonResponse({"logs": [], "material": None})

    logs = (
        InventoryUsageLog.objects.filter(raw_material_id=material_id)
        .select_related("raw_material")
        .order_by("-used_at")
    )
    material = logs.first().raw_material if logs.exists() else None

    data = [
        {
            "id": log.id,
            "quantity": str(log.quantity_used),
            "type": log.get_usage_type_display(),
            "type_key": log.usage_type,
            "reference": log.reference or "—",
            "note": log.note or "",
            "date": log.used_at.strftime("%Y/%m/%d %H:%M"),
        }
        for log in logs
    ]
    total = sum(float(d["quantity"]) for d in data)

    return JsonResponse(
        {
            "logs": data,
            "total_consumed": str(total),
            "material": (
                {
                    "name": material.name,
                    "unit": material.get_unit_display(),
                    "stock": str(material.quantity),
                }
                if material
                else None
            ),
        }
    )


# ══════════════════════════════════════════════════════════════════════════════
#  Warehouse Inventory — Ready Materials
# ══════════════════════════════════════════════════════════════════════════════


@staff_member_required
def ready_materials_page(request: HttpRequest):
    ready_materials = ReadyMaterial.objects.select_related('supplier', 'category').all()
    suppliers = Supplier.objects.all()
    raw_materials = RawMaterial.objects.all().order_by('name')
    categories = Category.objects.filter(is_active=True).order_by('order')
    return render(request, 'restaurant/ready_materials.html', {
        'ready_materials': ready_materials,
        'suppliers': suppliers,
        'unit_choices': ReadyMaterial.UNIT_CHOICES,
        'raw_materials': raw_materials,
        'categories': categories,
    })


@csrf_protect
@require_POST
def ready_material_save(request: HttpRequest):
    """ذخیره یا ویرایش ماده آماده — با کسر از ماده اولیه مرتبط."""
    try:
        pk = request.POST.get('id')
        name = request.POST.get('name', '').strip()
        if not name:
            return JsonResponse({'success': False, 'error': 'نام ماده الزامی است.'})

        description = request.POST.get('description', '').strip()
        unit = request.POST.get('unit', 'unit')
        quantity = Decimal(str(request.POST.get('quantity', 0) or 0))
        consume_quantity = Decimal(str(request.POST.get('consume_quantity', 0) or 0))
        purchase_price = int(float(request.POST.get('purchase_price', 0) or 0))
        selling_price = int(float(request.POST.get('selling_price', 0) or 0))
        minimum_stock = Decimal(str(request.POST.get('minimum_stock', 0) or 0))
        supplier_id = request.POST.get('supplier') or None
        barcode = request.POST.get('barcode', '').strip()
        raw_material_id = request.POST.get('raw_material_id') or None
        consume_quantity = Decimal(str(request.POST.get('consume_quantity', 0) or 0))
        category_id = request.POST.get('category') or None

        if purchase_price < 0:
            return JsonResponse({'success': False, 'error': 'قیمت خرید نمی‌تواند منفی باشد.'})
        if quantity < 0:
            return JsonResponse({'success': False, 'error': 'موجودی نمی‌تواند منفی باشد.'})

        raw_mat = None
        if raw_material_id and consume_quantity > 0:
            raw_mat = RawMaterial.objects.filter(pk=raw_material_id).first()
            if not raw_mat:
                return JsonResponse({'success': False, 'error': 'ماده اولیه یافت نشد.'})
            if raw_mat.quantity < consume_quantity:
                return JsonResponse({
                    'success': False,
                    'error': f'موجودی «{raw_mat.name}» ({raw_mat.quantity}) کمتر از مقدار مصرف ({consume_quantity}) است.'
                })

        if pk:
            mat = get_object_or_404(ReadyMaterial, pk=pk)
            old_raw = mat.source_raw_material
            old_consume = mat.consume_quantity or Decimal('0')
            if old_raw and old_consume > 0:
                old_raw.quantity += old_consume
                old_raw.save()

            mat.name = name
            mat.description = description
            mat.unit = unit
            mat.quantity = quantity
            mat.purchase_price = purchase_price
            mat.selling_price = selling_price
            mat.minimum_stock = minimum_stock
            mat.supplier_id = supplier_id
            mat.barcode = barcode
            mat.source_raw_material = raw_mat
            mat.consume_quantity = consume_quantity if raw_mat else Decimal('0')
            mat.category_id = category_id
            mat.save()
            if raw_mat and consume_quantity > 0:
                raw_mat.quantity -= consume_quantity
                raw_mat.save()
            msg = 'ویرایش شد.'
        else:
            mat = ReadyMaterial.objects.create(
                name=name, description=description, unit=unit,
                quantity=consume_quantity if raw_mat and consume_quantity > 0 else quantity,
                purchase_price=purchase_price,
                selling_price=selling_price, minimum_stock=minimum_stock,
                supplier_id=supplier_id, barcode=barcode,
                source_raw_material=raw_mat,
                consume_quantity=consume_quantity if raw_mat else Decimal('0'),
                category_id=category_id,
            )
            if raw_mat and consume_quantity > 0:
                raw_mat.quantity -= consume_quantity
                raw_mat.save()
            msg = 'اضافه شد.'

        return JsonResponse({
            'success': True, 'msg': msg,
            'item': {
                'id': mat.pk, 'name': mat.name,
                'description': mat.description or '',
                'unit': mat.unit, 'unit_display': mat.get_unit_display(),
                'quantity': float(mat.quantity),
                'purchase_price': int(mat.purchase_price),
                'selling_price': int(mat.selling_price),
                'minimum_stock': float(mat.minimum_stock),
                'supplier_id': mat.supplier_id,
                'supplier_name': mat.supplier.name if mat.supplier else '',
                'barcode': mat.barcode or '',
                'category_id': mat.category_id,
                'category_name': mat.category.name if mat.category else '',
                'total_value': int(mat.total_value),
                'stock_status': mat.stock_status,
            },
        })
    except Exception as exc:
        logger.exception('Error saving ready material')
        return JsonResponse({'success': False, 'error': str(exc)})

@csrf_protect
@require_POST
def ready_material_delete(request: HttpRequest):
    try:
        pk = request.POST.get("id")
        if not pk:
            return JsonResponse(
                {"success": False, "error": "شناسه ارسال نشد."}
            )
        mat = get_object_or_404(ReadyMaterial, pk=pk)
        name = mat.name

        if mat.source_raw_material and mat.consume_quantity > 0:
            raw = mat.source_raw_material
            raw.quantity += mat.consume_quantity
            raw.save()

        mat.delete()
        return JsonResponse({"success": True, "msg": f"«{name}» حذف شد."})
    except Exception as exc:
        logger.exception("Error deleting ready material")
        return JsonResponse({"success": False, "error": str(exc)})


@csrf_protect
@require_POST
def convert_to_ready_material(request: HttpRequest):
    """تبدیل ماده اولیه به ماده آماده."""
    try:
        raw_id = request.POST.get("raw_material_id")
        qty = float(request.POST.get("quantity", 0))
        selling_price = int(float(request.POST.get("selling_price", 0)))
        supplier_id = request.POST.get("supplier") or None

        if not raw_id:
            return JsonResponse(
                {"success": False, "error": "ماده اولیه انتخاب نشده."}
            )
        if qty <= 0:
            return JsonResponse(
                {"success": False, "error": "مقدار باید بیشتر از صفر باشد."}
            )

        raw_mat = get_object_or_404(RawMaterial, pk=raw_id)

        if qty > float(raw_mat.quantity):
            return JsonResponse(
                {
                    "success": False,
                    "error": f"موجودی کافی نیست. حداکثر: {int(raw_mat.quantity)}",
                }
            )

        raw_mat.quantity -= Decimal(str(qty))
        raw_mat.save(update_fields=["quantity"])

        ready = ReadyMaterial.objects.create(
            name=raw_mat.name,
            description="تبدیل شده از ماده اولیه",
            unit=raw_mat.unit,
            quantity=qty,
            purchase_price=int(raw_mat.price),
            selling_price=selling_price,
            minimum_stock=0,
            supplier_id=supplier_id,
        )

        return JsonResponse(
            {
                "success": True,
                "msg": f"«{raw_mat.name}» به مواد آماده اضافه شد.",
                "item": {
                    "id": ready.pk,
                    "name": ready.name,
                    "description": ready.description or "",
                    "unit": ready.unit,
                    "unit_display": ready.get_unit_display(),
                    "quantity": float(ready.quantity),
                    "purchase_price": int(ready.purchase_price),
                    "selling_price": int(ready.selling_price),
                    "minimum_stock": float(ready.minimum_stock),
                    "supplier_id": ready.supplier_id,
                    "supplier_name": (
                        ready.supplier.name if ready.supplier else ""
                    ),
                    "barcode": ready.barcode or "",
                    "total_value": int(ready.total_value),
                    "stock_status": ready.stock_status,
                },
            }
        )
    except Exception as exc:
        logger.exception("Error converting to ready material")
        return JsonResponse({"success": False, "error": str(exc)})

@csrf_protect
@require_POST
@staff_member_required
def ready_material_update_price(request: HttpRequest):
    """بروزرسانی قیمت فروش ماده آماده از آشپزخانه"""
    try:
        data = json_module.loads(request.body)
        rm_id = data.get('id')
        selling_price = int(data.get('selling_price', 0))

        if not rm_id:
            return JsonResponse({'success': False, 'error': 'شناسه ارسال نشد.'})
        if selling_price < 0:
            return JsonResponse({'success': False, 'error': 'قیمت نمی‌تواند منفی باشد.'})

        rm = get_object_or_404(ReadyMaterial, pk=rm_id)
        rm.selling_price = selling_price
        rm.save(update_fields=['selling_price'])

        return JsonResponse({
            'success': True,
            'msg': f'قیمت «{rm.name}» بروزرسانی شد.',
            'item': {
                'id': rm.pk,
                'name': rm.name,
                'selling_price': int(rm.selling_price),
                'purchase_price': int(rm.purchase_price),
            }
        })
    except Exception as exc:
        logger.exception('Error updating ready material price')
        return JsonResponse({'success': False, 'error': str(exc)})

# ══════════════════════════════════════════════════════════════════════════════
#  KITCHEN MANAGEMENT — VIEWS
# ══════════════════════════════════════════════════════════════════════════════


@staff_member_required
def kitchen_page(request: HttpRequest):
    recipes = list(Recipe.objects.values("id").annotate(name=F("food__name")))
    categories = list(KitchenProduct.CATEGORY_CHOICES)

    foods_list = []
    for f in Food.objects.select_related('category').all().order_by('category__order', 'name'):
        foods_list.append({
            'id': f.id,
            'name': f.name,
            'category_id': f.category_id,
            'category_name': f.category.name,
            'final_price': int(f.final_price),
            'image': f.image.url if f.image else '',
            'is_ready': False,
            'purchase_price': 0,
        })

    food_cats = list(
        Category.objects.filter(is_active=True).order_by('order').values('id', 'name')
    )
    existing_names = {c['name'] for c in food_cats}

    ready_materials = ReadyMaterial.objects.filter(
        quantity__gt=0
    ).exclude(category__isnull=True).select_related('category')

    for rm in ready_materials:
        foods_list.append({
            'id': f'ready_{rm.id}',
            'name': rm.name,
            'category_id': rm.category_id,
            'category_name': rm.category.name,
            'final_price': int(rm.selling_price or 0),
            'image': '',
            'is_ready': True,
            'purchase_price': int(rm.purchase_price or 0),
        })
        if rm.category.name not in existing_names:
            food_cats.append({'id': rm.category_id, 'name': rm.category.name})
            existing_names.add(rm.category.name)

    return render(request, "restaurant/kitchen_page.html", {
        "recipes_json": json_module.dumps(recipes, ensure_ascii=False),
        "categories_json": json_module.dumps(categories, ensure_ascii=False),
        "foods_json": json_module.dumps(foods_list, ensure_ascii=False),
        "food_cats_json": json_module.dumps(food_cats, ensure_ascii=False),
    })


@api_view(["GET"])
@staff_member_required
def kitchen_dashboard_api(request):
    return JsonResponse(generate_kitchen_dashboard(), safe=False)


class KitchenProductListCreate(generics.ListCreateAPIView):
    serializer_class = KitchenProductSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = KitchenProduct.objects.select_related("recipe").prefetch_related(
            "discounts"
        ).all()
        cat = self.request.query_params.get("category")
        if cat:
            qs = qs.filter(category=cat)
        active = self.request.query_params.get("active")
        if active is not None:
            qs = qs.filter(is_active=active.lower() == "true")
        return qs


class KitchenProductDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = KitchenProduct.objects.select_related("recipe").all()
    serializer_class = KitchenProductSerializer
    permission_classes = [permissions.IsAuthenticated]


@api_view(["GET"])
@permission_classes([permissions.IsAuthenticated])
def kitchen_product_capacity(request, pk: int):
    try:
        product = KitchenProduct.objects.select_related("recipe").get(pk=pk)
    except KitchenProduct.DoesNotExist:
        return Response(
            {"error": "محصول یافت نشد."}, status=status.HTTP_404_NOT_FOUND
        )
    mx, lim = calculate_max_production(product)
    req = get_required_materials(product, 1)
    return Response(
        {
            "product_id": product.id,
            "product_name": product.name,
            "max_production": mx,
            "limiting_material": lim,
            "required_per_unit": req,
        }
    )

@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])  # ★FIX
def kitchen_product_produce(request, pk: int):
    try:
        product = KitchenProduct.objects.select_related("recipe").get(pk=pk)
    except KitchenProduct.DoesNotExist:
        return Response(
            {"error": "محصول یافت نشد."}, status=status.HTTP_404_NOT_FOUND
        )

    try:
        quantity = int(request.data.get("quantity", 0))
    except (TypeError, ValueError):
        return Response(
            {"error": "تعداد نامعتبر است."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    notes = request.data.get("notes", "")
    if quantity <= 0:
        return Response(
            {"error": "تعداد باید بیشتر از صفر باشد."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        batch = produce_item(
            kitchen_product=product,
            quantity=quantity,
            user=request.user,
            notes=notes,
        )
        return Response(
            {
                "success": True,
                "msg": f'{quantity} واحد از «{product.name}» تولید شد.',
                "batch_id": batch.id,
                "production_cost": batch.production_cost,
            }
        )
    except ValidationError as e:
        msgs = e.messages if hasattr(e, "messages") else [str(e)]
        return Response(
            {"error": msgs}, status=status.HTTP_400_BAD_REQUEST
        )

class KitchenInventoryList(generics.ListAPIView):
    queryset = KitchenInventory.objects.select_related("kitchen_product").all()
    serializer_class = KitchenInventorySerializer
    permission_classes = [permissions.IsAuthenticated]


class ProductionPlanListCreate(generics.ListCreateAPIView):
    serializer_class = ProductionPlanSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return (
            ProductionPlan.objects.prefetch_related("items__kitchen_product")
            .select_related("created_by")
            .order_by("-date", "-created_at")
        )

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)


class ProductionPlanDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = ProductionPlan.objects.prefetch_related(
        "items__kitchen_product"
    ).all()
    serializer_class = ProductionPlanSerializer
    permission_classes = [permissions.IsAuthenticated]


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def kitchen_calculate_materials(request):
    """محاسبه کل مواد اولیه و نیمه‌آماده مورد نیاز برای لیست محصولات."""
    items = request.data.get("items", [])
    if not items:
        return Response(
            {"error": "آیتمی ارسال نشده."}, status=status.HTTP_400_BAD_REQUEST
        )

    products_summary = []
    materials_map: dict[tuple, dict] = {}

    for item in items:
        pid = item.get("product_id")
        qty = int(item.get("quantity", 0))
        if not pid or qty <= 0:
            continue
        try:
            kp = KitchenProduct.objects.select_related("recipe").get(pk=pid)
        except KitchenProduct.DoesNotExist:
            continue

        products_summary.append({"name": kp.name, "quantity": qty})

        reqs = get_required_materials(kp, qty)
        for r in reqs:
            key = (r["type"], r["id"])
            if key not in materials_map:
                materials_map[key] = {
                    "name": r["name"],
                    "type": r["type"],
                    "required": 0,
                    "available": r["available"],
                    "unit": r["unit_display"],
                }
            materials_map[key]["required"] += r["total_needed"]

    raw_materials = []
    semi_materials = []
    shortage_count = 0

    for m in materials_map.values():
        m["required"] = round(m["required"], 2)
        if m["available"] < m["required"]:
            shortage_count += 1
        entry = {
            "name": m["name"],
            "required": m["required"],
            "available": round(m["available"], 2),
            "unit": m["unit"],
        }
        if m["type"] == "raw_material":
            raw_materials.append(entry)
        else:
            semi_materials.append(entry)

    return Response(
        {
            "products": products_summary,
            "raw_materials": raw_materials,
            "semi_materials": semi_materials,
            "shortage_count": shortage_count,
        }
    )


@api_view(["POST"])
@permission_classes([permissions.IsAuthenticated])
def production_plan_approve(request, pk: int):
    try:
        plan = ProductionPlan.objects.get(pk=pk)
    except ProductionPlan.DoesNotExist:
        return Response(
            {"error": "برنامه یافت نشد."}, status=status.HTTP_404_NOT_FOUND
        )
    try:
        approve_production_plan(plan, user=request.user)
        return Response({"success": True, "msg": "برنامه تأیید شد."})
    except ValidationError as e:
        msgs = e.messages if hasattr(e, "messages") else [str(e)]
        return Response(
            {"error": msgs}, status=status.HTTP_400_BAD_REQUEST
        )
    except Exception as e:
        logger.exception("Error approving plan %s", pk)
        return Response(
            {"error": f"خطای سرور: {e}"},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(["POST"])
@permission_classes([IsOwnerOrManagerOrKitchenStaff])
def production_plan_execute(request, pk: int):
    try:
        plan = ProductionPlan.objects.prefetch_related(
            "items__kitchen_product"
        ).get(pk=pk)
    except ProductionPlan.DoesNotExist:
        return Response(
            {"error": "برنامه یافت نشد."}, status=status.HTTP_404_NOT_FOUND
        )
    try:
        batches = execute_production_plan(plan, user=request.user)
        return Response(
            {
                "success": True,
                "msg": f"برنامه اجرا شد — {len(batches)} محصول تولید شد.",
                "batch_ids": [b.id for b in batches],
            }
        )
    except ValidationError as e:
        msgs = e.messages if hasattr(e, "messages") else [str(e)]
        return Response(
            {"error": msgs}, status=status.HTTP_400_BAD_REQUEST
        )


class KitchenDiscountListCreate(generics.ListCreateAPIView):
    serializer_class = KitchenDiscountSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = KitchenDiscount.objects.select_related("kitchen_product").all()
        active = self.request.query_params.get("active")
        if active is not None:
            qs = qs.filter(is_active=active.lower() == "true")
        return qs


class KitchenDiscountDetail(generics.RetrieveUpdateDestroyAPIView):
    queryset = KitchenDiscount.objects.select_related("kitchen_product").all()
    serializer_class = KitchenDiscountSerializer
    permission_classes = [permissions.IsAuthenticated]


class ProductionLogList(generics.ListAPIView):
    serializer_class = ProductionLogSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = ProductionLog.objects.select_related("kitchen_product", "user")
        action_filter = self.request.query_params.get("action")
        if action_filter:
            qs = qs.filter(action=action_filter)
        product_id = self.request.query_params.get("product")
        if product_id:
            qs = qs.filter(kitchen_product_id=product_id)
        return qs.order_by("-created_at")[:100]


# ══════════════════════════════════════════════════════════════════════════════
#  Food & Category Management
# ══════════════════════════════════════════════════════════════════════════════


@staff_member_required
def food_management_page(request: HttpRequest):
    foods_data, categories_data = _build_foods_with_discounts()
    return render(
        request,
        "restaurant/food_management.html",
        {
            "foods_json": json_module.dumps(foods_data, ensure_ascii=False),
            "categories_json": json_module.dumps(
                categories_data, ensure_ascii=False
            ),
        },
    )


@staff_member_required
def food_management_api(request: HttpRequest):
    """GET — لیست زنده غذاها با تخفیف‌ها برای صفحه مدیریت."""
    foods_data, _ = _build_foods_with_discounts()
    return JsonResponse({"foods": foods_data})


@csrf_protect
@require_POST
def food_save(request: HttpRequest):
    try:
        pk = request.POST.get("id")
        name = request.POST.get("name", "").strip()
        category_id = request.POST.get("category")
        final_price = int(float(request.POST.get("final_price", 0)))

        if not name:
            return JsonResponse(
                {"success": False, "error": "نام غذا الزامی است."}
            )
        if not category_id:
            return JsonResponse(
                {"success": False, "error": "دسته‌بندی الزامی است."}
            )
        if final_price < 0:
            return JsonResponse(
                {"success": False, "error": "قیمت نمی‌تواند منفی باشد."}
            )

        if pk:
            food = get_object_or_404(Food, pk=pk)
            food.name = name
            food.category_id = category_id
            food.final_price = final_price
            if "image" in request.FILES:
                food.image = request.FILES["image"]
            food.save()
            msg = "ویرایش شد."
        else:
            food = Food.objects.create(
                name=name,
                category_id=category_id,
                final_price=final_price,
            )
            if "image" in request.FILES:
                food.image = request.FILES["image"]
                food.save()
            msg = "اضافه شد."

        return JsonResponse(
            {
                "success": True,
                "msg": msg,
                "item": {
                    "id": food.pk,
                    "name": food.name,
                    "category_id": food.category_id,
                    "category_name": food.category.name,
                    "final_price": int(food.final_price),
                    "image": food.image.url if food.image else "",
                },
            }
        )
    except Exception as exc:
        logger.exception("Error saving food")
        return JsonResponse({"success": False, "error": str(exc)})


@csrf_protect
@require_POST
def food_delete(request: HttpRequest):
    try:
        pk = request.POST.get("id")
        if not pk:
            return JsonResponse(
                {"success": False, "error": "شناسه ارسال نشد."}
            )
        food = get_object_or_404(Food, pk=pk)
        name = food.name
        food.delete()
        return JsonResponse({"success": True, "msg": f"«{name}» حذف شد."})
    except Exception as exc:
        logger.exception("Error deleting food")
        return JsonResponse({"success": False, "error": str(exc)})


@csrf_protect
@require_POST
def category_save(request: HttpRequest):
    try:
        pk = request.POST.get("id")
        name = request.POST.get("name", "").strip()
        order = int(request.POST.get("order", 0))

        if not name:
            return JsonResponse(
                {"success": False, "error": "نام دسته‌بندی الزامی است."}
            )

        if pk:
            cat = get_object_or_404(Category, pk=pk)
            cat.name = name
            cat.order = order
            if "image" in request.FILES:
                cat.image = request.FILES["image"]
            cat.save()
            msg = "ویرایش شد."
        else:
            cat = Category.objects.create(name=name, order=order)
            if "image" in request.FILES:
                cat.image = request.FILES["image"]
                cat.save()
            msg = "اضافه شد."

        return JsonResponse(
            {
                "success": True,
                "msg": msg,
                "item": {
                    "id": cat.pk,
                    "name": cat.name,
                    "order": cat.order,
                    "image": cat.image.url if cat.image else "",
                },
            }
        )
    except Exception as exc:
        logger.exception("Error saving category")
        return JsonResponse({"success": False, "error": str(exc)})


@csrf_protect
@require_POST
def category_delete(request: HttpRequest):
    try:
        pk = request.POST.get("id")
        if not pk:
            return JsonResponse(
                {"success": False, "error": "شناسه ارسال نشد."}
            )
        cat = get_object_or_404(Category, pk=pk)
        if cat.food_set.exists():
            return JsonResponse(
                {
                    "success": False,
                    "error": (
                        f"دسته‌بندی «{cat.name}» دارای غذا است "
                        f"و قابل حذف نیست."
                    ),
                }
            )
        name = cat.name
        cat.delete()
        return JsonResponse({"success": True, "msg": f"«{name}» حذف شد."})
    except Exception as exc:
        logger.exception("Error deleting category")
        return JsonResponse({"success": False, "error": str(exc)})


# ══════════════════════════════════════════════════════════════════════════════
#  POS — صندوق فروش
# ══════════════════════════════════════════════════════════════════════════════

_DRINK_KEYWORDS = [
    "نوشابه", "دلستر", "دوغ", "ماءالشعیر", "ماء الشعیر", "آب معدنی",
]


def _is_drink(name: str) -> bool:
    """آیا نام ماده آماده شبیه نوشیدنی است؟"""
    return any(k in name for k in _DRINK_KEYWORDS)


@staff_member_required
def pos_page(request: HttpRequest):
    """صفحه صندوق فروش — غذاها + مواد آماده در دسته‌بندی مربوطه."""
    foods_data, cats_data = _build_foods_with_discounts()
    existing_names = {c['name']: c['id'] for c in cats_data}

    # ═══ اضافه کردن موجودی به غذاهای آشپزخانه ═══
    for food_item in foods_data:
        if food_item.get('is_ready'):
            continue
        stock = 0
        kitchen_product = KitchenProduct.objects.filter(name=food_item['name']).first()
        if kitchen_product:
            try:
                inv = kitchen_product.get_inventory()
                if inv:
                    stock = inv.available_quantity or 0
            except Exception:
                stock = 0
        food_item['stock'] = stock

    # ═══ مواد آماده ═══
    ready_materials = ReadyMaterial.objects.filter(
        quantity__gt=0
    ).select_related('category')

    for rm in ready_materials:
        if rm.category:
            cat_id = rm.category_id
            cat_name = rm.category.name
        else:
            cat_name = 'سایر'
            cat_id = existing_names.get('سایر', -99)

        foods_data.append({
            'id': f'ready_{rm.id}',
            'name': rm.name,
            'category_id': cat_id,
            'category_name': cat_name,
            'final_price': int(rm.selling_price or 0),
            'kitchen_price': int(rm.selling_price or 0),
            'has_kitchen': False,
            'discount': None,
            'image': '',
            'is_ready': True,
            'stock': int(rm.quantity or 0),
        })

        if rm.category and rm.category.name not in existing_names:
            cats_data.append({'id': rm.category_id, 'name': rm.category.name})
            existing_names[rm.category.name] = rm.category_id

    if any(not rm.category for rm in ready_materials) and 'سایر' not in existing_names:
        cats_data.append({'id': -99, 'name': 'سایر'})

    return render(request, 'restaurant/pos.html', {
        'foods_json': json_module.dumps(foods_data, ensure_ascii=False),
        'categories_json': json_module.dumps(cats_data, ensure_ascii=False),
    })

@csrf_protect
@require_POST
@staff_member_required
def pos_create_order(request: HttpRequest):
    """ثبت سفارش از صفحه صندوق — شامل نوشیدنی‌های آماده + بررسی موجودی."""
    try:
        data = json_module.loads(request.body)
        customer_name = data.get("customer_name", "").strip()
        phone = data.get("phone", "").strip()
        items = data.get("items", [])

        if not items:
            return JsonResponse({"success": False, "error": "هیچ غذایی انتخاب نشده"})

        # ═══ اعتبارسنجی + بررسی موجودی ═══
        validated_items = []
        stock_errors = []

        for item in items:
            qty = int(item.get("quantity", 1))
            if qty <= 0:
                continue

            raw_id = item.get("food_id") or item.get("id")
            is_ready = item.get("is_ready", False)

            # ── مواد آماده ──
            if is_ready or (isinstance(raw_id, str) and raw_id.startswith("ready_")):
                rm_id = int(str(raw_id).replace("ready_", ""))
                rm = ReadyMaterial.objects.filter(id=rm_id).first()
                if not rm:
                    return JsonResponse({
                        "success": False,
                        "error": f"کالای آماده با شناسه {rm_id} پیدا نشد"
                    })
                if qty > int(rm.quantity):
                    stock_errors.append(
                        f"{rm.name}: سفارش {qty} ولی موجودی {int(rm.quantity)}"
                    )
                    continue
                validated_items.append({
                    "type": "ready",
                    "obj": rm,
                    "qty": qty,
                    "price": int(rm.selling_price),
                })

            # ── غذای آشپزخانه ──
            else:
                food = Food.objects.filter(id=int(raw_id)).first()
                if not food:
                    return JsonResponse({
                        "success": False,
                        "error": f"غذا با شناسه {raw_id} پیدا نشد"
                    })

                db_price = int(food.final_price)

                # پیدا کردن محصول آشپزخانه
                kp = None
                if hasattr(food, "recipe") and food.recipe:
                    kp = food.recipe.kitchen_products.first()
                if not kp:
                    kp = KitchenProduct.objects.filter(name=food.name).first()

                # بررسی موجودی
                if kp:
                    inv = kp.get_inventory()
                    available = inv.available_quantity if inv else 0
                    if qty > available:
                        stock_errors.append(
                            f"{food.name}: سفارش {qty} ولی موجودی {available}"
                        )
                        continue

                validated_items.append({
                    "type": "food",
                    "obj": food,
                    "kp": kp,
                    "qty": qty,
                    "price": db_price,
                })

        # ═══ اگه موجودی کافی نبود ═══
        if stock_errors:
            return JsonResponse({
                "success": False,
                "error": "موجودی کافی نیست: " + " | ".join(stock_errors)
            })

        if not validated_items:
            return JsonResponse({
                "success": False,
                "error": "هیچ آیتم معتبری وجود ندارد"
            })

        # ═══ ثبت سفارش ═══
        with transaction.atomic():
            order = Order.objects.create(
                customer_name=customer_name or "مشتری",
                phone=phone,
                status="pending",
                total_price=0,
            )

            total = 0
            order_items = []

            for vi in validated_items:
                qty = vi["qty"]
                price = vi["price"]
                line_total = price * qty
                total += line_total

                # ── مواد آماده: کسر از موجودی ──
                if vi["type"] == "ready":
                    rm = vi["obj"]
                    rm.quantity -= qty
                    rm.save(update_fields=["quantity"])
                    OrderItem.objects.create(
                        order=order, food=None, quantity=qty, price=price
                    )
                    order_items.append({
                        "name": rm.name,
                        "quantity": qty,
                        "price": price,
                        "line_total": line_total,
                    })

                # ── غذا: کسر از موجودی آشپزخانه ──
                else:
                    food = vi["obj"]
                    kp = vi["kp"]
                    if kp:
                        inv = kp.get_inventory()
                        if inv:
                            inv.quantity -= qty
                            inv.save(update_fields=["quantity", "updated_at"])
                    OrderItem.objects.create(
                        order=order, food=food, quantity=qty, price=price
                    )
                    order_items.append({
                        "name": food.name,
                        "quantity": qty,
                        "price": price,
                        "line_total": line_total,
                    })

            order.total_price = total
            order.save()

        return JsonResponse({
            "success": True,
            "order_id": order.id,
            "customer_name": order.customer_name,
            "total_price": total,
            "items": order_items,
            "created_at": order.created_at.strftime("%Y-%m-%d %H:%M"),
            "msg": f"سفارش #{order.id} ثبت شد",
        })

    except Exception as exc:
        logger.exception("Error creating POS order")
        return JsonResponse({"success": False, "error": str(exc)})

@staff_member_required
def pos_receipt(request: HttpRequest, pk: int):
    order = get_object_or_404(Order, pk=pk)
    items_qs = order.items.select_related("food").all()

    items = []
    for item in items_qs:
        price = int(item.price)
        qty = item.quantity
        items.append({
            "food_name": item.food.name if item.food else "—",
            "quantity": qty,
            "price": price,
            "line_total": price * qty,
        })

    discount_amount = 0
    if hasattr(order, "discount_amount") and order.discount_amount:
        discount_amount = int(order.discount_amount)

    final_amount = int(order.total_price) - discount_amount

    return render(request, "restaurant/receipt.html", {
        "order": order,
        "items": items,
        "discount_amount": discount_amount if discount_amount > 0 else None,
        "final_amount": final_amount,
        "payment_method": getattr(order, "payment_method", None),
        "trace_number": getattr(order, "trace_number", None),
        "restaurant_name": getattr(settings, "RESTAURANT_NAME", "رستوران"),
        "restaurant_phone": getattr(settings, "RESTAURANT_PHONE", ""),
        "restaurant_address": getattr(settings, "RESTAURANT_ADDRESS", ""),
    })

@require_POST
@login_required
def pos_validate_coupon(request):
    try:
        data = json.loads(request.body)
        code = (data.get('code') or '').strip().upper()
        subtotal = int(data.get('subtotal') or 0)
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({'success': False, 'error': 'داده نامعتبر'})

    if not code:
        return JsonResponse({'success': False, 'error': 'کد تخفیف وارد نشده'})

    try:
        coupon = Coupon.objects.get(code__iexact=code)
    except Coupon.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'کد تخفیف نامعتبر است'})

    # بررسی اعتبار
    if not coupon.is_valid:
        return JsonResponse({'success': False, 'error': 'این کد منقضی شده یا غیرفعال است'})

    # حداقل مبلغ سفارش
    if coupon.min_order_amount and subtotal < coupon.min_order_amount:
        return JsonResponse({
            'success': False,
            'error': 'حداقل مبلغ سفارش برای این کد: ' + str(coupon.min_order_amount) + ' تومان'
        })

    # محاسبه تخفیف
    from decimal import Decimal
    discount = coupon.calculate_discount(Decimal(str(subtotal)))

    # افزایش شمارنده

    # ساخت توضیح
    if coupon.description:
        desc = coupon.description
    elif coupon.discount_type == 'percentage':
        desc = str(coupon.discount_value) + '% تخفیف'
        if coupon.max_discount_amount:
            desc += ' (سقف ' + str(coupon.max_discount_amount) + ' تومان)'
    else:
        desc = str(coupon.discount_value) + ' تومان تخفیف'

    return JsonResponse({
        'success': True,
        'discount_type': coupon.discount_type,
        'value': int(discount),  # مبلغ واقعی تخفیف محاسبه‌شده
        'description': desc
    })
# ══════════════════════════════════════════════════════════════════════════════
#  گزارش روزانه + بستن روز + ضایعات
# ══════════════════════════════════════════════════════════════════════════════

@staff_member_required
def pos_daily_report(request: HttpRequest):
    """گزارش فروش روزانه — آمار + پرفروش‌ها + لیست سفارشات."""
    try:
        date_str = request.GET.get('date', '')
        if date_str:
            target_date = datetime.date.fromisoformat(date_str)
        else:
            target_date = timezone.localdate()

        start = timezone.make_aware(
            timezone.datetime.combine(target_date, timezone.datetime.min.time())
        )
        end = timezone.make_aware(
            timezone.datetime.combine(target_date, timezone.datetime.max.time())
        )

        orders = Order.objects.filter(
            created_at__range=(start, end)
        ).prefetch_related('items__food')

        order_count = orders.count()
        total_sales = sum(o.total_price for o in orders)

        # پرفروش‌ترین‌ها
        top_items = (
            OrderItem.objects.filter(order__in=orders, food__isnull=False)
            .values('food__name')
            .annotate(qty=Sum('quantity'), total=Sum('price'))
            .order_by('-qty')[:10]
        )
        top_list = [
            {'name': t['food__name'], 'qty': t['qty'], 'total': int(t['total'] or 0)}
            for t in top_items
        ]

        # لیست سفارشات
        orders_list = []
        for o in orders.order_by('-created_at'):
            orders_list.append({
                'id': o.id,
                'customer': o.customer_name,
                'items_count': o.items.count(),
                'total': int(o.total_price),
                'time': o.created_at.strftime('%H:%M'),
            })

        # ضایعات امروز
        waste_logs = WasteLog.objects.filter(created_at__range=(start, end))
        waste_total = sum(w.quantity for w in waste_logs)

        # تخفیف‌ها
        discount_total = 0
        for o in orders:
            for item in o.items.select_related('food').all():
                if item.food:
                    kp = None
                    if hasattr(item.food, 'recipe') and item.food.recipe:
                        kp = item.food.recipe.kitchen_products.first()
                    if not kp:
                        kp = KitchenProduct.objects.filter(name=item.food.name).first()
                    if kp:
                        kitchen_price = int(kp.selling_price)  # ★FIX: قیمت آشپزخانه
                        di = _get_food_discount_info(kp)
                        if di:
                            # ★FIX: بجای di['original_price'] که وجود نداره
                            discount_total += (
                                kitchen_price - int(di['discounted_price'])
                            ) * item.quantity

        return JsonResponse({
            'success': True,
            'total_sales': int(total_sales),
            'order_count': order_count,
            'discount_total': discount_total,
            'waste_total': waste_total,
            'top_items': top_list,
            'orders': orders_list,
        })
    except Exception as exc:
        logger.exception("Error in daily report")
        return JsonResponse({'success': False, 'error': str(exc)})

class MembershipLevelViewSet(viewsets.ModelViewSet):
    queryset = MembershipLevel.objects.all()
    serializer_class = MembershipLevelSerializer

    @action(detail=False, methods=["post"], url_path="seed")
    def seed(self, request):
        return Response({"message": seed_membership_levels()})


class CustomerViewSet(viewsets.ModelViewSet):

    def get_queryset(self):
        qs = CustomerProfile.objects.select_related("membership_level").all()
        search = self.request.query_params.get("search")
        if search:
            qs = qs.filter(
                Q(phone__icontains=search)
                | Q(first_name__icontains=search)
                | Q(last_name__icontains=search)
            )
        level = self.request.query_params.get("level")
        if level:
            qs = qs.filter(membership_level__name=level)
        is_active = self.request.query_params.get("active")
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == "true")
        return qs

    def get_serializer_class(self):
        match self.action:
            case "list":
                return CustomerListSerializer
            case "create":
                return CustomerCreateSerializer
            case "update" | "partial_update":
                return CustomerUpdateSerializer
            case _:
                return CustomerDetailSerializer

    def create(self, request, *args, **kwargs):
        ser = CustomerCreateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        result = register_customer(**ser.validated_data)
        if not result["success"]:
            return Response(
                {"error": result["error"]}, status=status.HTTP_400_BAD_REQUEST
            )
        return Response(
            CustomerDetailSerializer(
                result["customer"], context={"request": request}
            ).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"], url_path="earn-points")
    def earn_points(self, request, pk=None):
        customer = self.get_object()
        ser = EarnPointsSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        result = earn_points_for_order(
            customer=customer, **ser.validated_data
        )
        return Response(
            result,
            status=(
                status.HTTP_200_OK
                if result["success"]
                else status.HTTP_400_BAD_REQUEST
            ),
        )

    @action(detail=True, methods=["post"], url_path="redeem-points")
    def redeem_points_action(self, request, pk=None):
        customer = self.get_object()
        ser = RedeemPointsSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        result = redeem_points(
            customer=customer,
            points=ser.validated_data["points"],
            order_id=ser.validated_data.get("order_id"),
        )
        return Response(
            result,
            status=(
                status.HTTP_200_OK
                if result["success"]
                else status.HTTP_400_BAD_REQUEST
            ),
        )

    @action(detail=True, methods=["get"], url_path="wallet")
    def wallet(self, request, pk=None):
        customer = self.get_object()
        wallet_obj = LoyaltyWallet.objects.filter(customer=customer).first()
        if not wallet_obj:
            return Response({"balance": 0, "transactions": []})
        txns = wallet_obj.transactions.all()[:20]
        return Response(
            {
                "wallet": WalletSerializer(wallet_obj).data,
                "transactions": WalletTransactionSerializer(
                    txns, many=True
                ).data,
            }
        )

    @action(detail=True, methods=["post"], url_path="wallet/deposit")
    def wallet_deposit_action(self, request, pk=None):
        customer = self.get_object()
        ser = WalletDepositSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        result = wallet_deposit(
            customer=customer,
            amount=ser.validated_data["amount"],
            description=ser.validated_data.get("description", ""),
        )
        return Response(
            result,
            status=(
                status.HTTP_200_OK
                if result["success"]
                else status.HTTP_400_BAD_REQUEST
            ),
        )

    @action(detail=True, methods=["post"], url_path="wallet/debit")
    def wallet_debit_action(self, request, pk=None):
        customer = self.get_object()
        ser = WalletDebitSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        result = wallet_debit(
            customer=customer,
            amount=ser.validated_data["amount"],
            description=ser.validated_data.get("description", ""),
            order_id=ser.validated_data.get("order_id"),
        )
        return Response(
            result,
            status=(
                status.HTTP_200_OK
                if result["success"]
                else status.HTTP_400_BAD_REQUEST
            ),
        )

    @action(detail=True, methods=["post"], url_path="validate-coupon")
    def validate_coupon_action(self, request, pk=None):
        customer = self.get_object()
        ser = CouponValidateSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        result = validate_coupon(
            code=ser.validated_data["code"],
            customer=customer,
            order_amount=ser.validated_data["order_amount"],
        )
        if result.get("coupon"):
            result["coupon"] = CouponDetailSerializer(result["coupon"]).data
        return Response(result)

    @action(detail=True, methods=["post"], url_path="apply-coupon")
    def apply_coupon_action(self, request, pk=None):
        customer = self.get_object()
        ser = CouponApplySerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        result = apply_coupon(
            code=ser.validated_data["code"],
            customer=customer,
            order_amount=ser.validated_data["order_amount"],
            order_id=ser.validated_data.get("order_id"),
        )
        return Response(
            result,
            status=(
                status.HTTP_200_OK
                if result["success"]
                else status.HTTP_400_BAD_REQUEST
            ),
        )

    @action(detail=True, methods=["get"], url_path="redemptions")
    def redemptions(self, request, pk=None):
        customer = self.get_object()
        qs = customer.reward_redemptions.select_related("reward").all()[:20]
        return Response(RewardRedemptionSerializer(qs, many=True).data)

    @action(detail=True, methods=["get"], url_path="transactions")
    def transactions(self, request, pk=None):
        customer = self.get_object()
        return Response(
            LoyaltyTransactionSerializer(
                customer.loyalty_transactions.all()[:50], many=True
            ).data
        )

    @action(detail=True, methods=["get"], url_path="notifications")
    def notifications(self, request, pk=None):
        customer = self.get_object()
        return Response(
            {
                "notifications": NotificationSerializer(
                    customer.notifications.all()[:30], many=True
                ).data,
                "unread_count": customer.notifications.filter(
                    is_read=False
                ).count(),
            }
        )

    @action(detail=True, methods=["post"], url_path="check-birthday")
    def check_birthday(self, request, pk=None):
        return Response(check_and_grant_birthday_bonus(self.get_object()))

    @action(detail=True, methods=["post"], url_path="check-level")
    def check_level(self, request, pk=None):
        result = check_level_upgrade(self.get_object())
        return Response(
            {
                "upgraded": result["upgraded"],
                "new_level": (
                    MembershipLevelSerializer(result["new_level"]).data
                    if result["new_level"]
                    else None
                ),
                "current_level": (
                    MembershipLevelSerializer(result["current_level"]).data
                    if result["current_level"]
                    else None
                ),
            }
        )


class CouponViewSet(viewsets.ModelViewSet):

    def get_queryset(self):
        qs = Coupon.objects.prefetch_related("applicable_levels").all()
        is_active = self.request.query_params.get("active")
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == "true")
        coupon_type = self.request.query_params.get("type")
        if coupon_type:
            qs = qs.filter(coupon_type=coupon_type)
        search = self.request.query_params.get("search")
        if search:
            qs = qs.filter(
                Q(code__icontains=search) | Q(name__icontains=search)
            )
        return qs

    def get_serializer_class(self):
        match self.action:
            case "list":
                return CouponListSerializer
            case "create" | "update" | "partial_update":
                return CouponCreateSerializer
            case _:
                return CouponDetailSerializer

    @action(detail=True, methods=["post"], url_path="toggle-active")
    def toggle_active(self, request, pk=None):
        coupon = self.get_object()
        coupon.is_active = not coupon.is_active
        coupon.save(update_fields=["is_active"])
        return Response({"is_active": coupon.is_active})


class RewardViewSet(viewsets.ModelViewSet):

    def get_queryset(self):
        qs = Reward.objects.select_related("min_membership_level").all()
        is_active = self.request.query_params.get("active")
        if is_active is not None:
            qs = qs.filter(is_active=is_active.lower() == "true")
        category = self.request.query_params.get("category")
        if category:
            qs = qs.filter(category=category)
        return qs

    def get_serializer_class(self):
        match self.action:
            case "list":
                return RewardListSerializer
            case "create" | "update" | "partial_update":
                return RewardCreateSerializer
            case _:
                return RewardDetailSerializer

    @action(detail=True, methods=["post"], url_path="redeem")
    def redeem_action(self, request, pk=None):
        phone = (
            request.data.get("phone")
            or request.headers.get("X-Customer-Phone")
        )
        if not phone:
            return Response(
                {"error": "شماره موبایل لازم است."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        customer = CustomerProfile.objects.filter(phone=phone).first()
        if not customer:
            return Response(
                {"error": "مشتری یافت نشد."},
                status=status.HTTP_404_NOT_FOUND,
            )
        result = redeem_reward(customer=customer, reward_id=pk)
        return Response(
            result,
            status=(
                status.HTTP_200_OK
                if result["success"]
                else status.HTTP_400_BAD_REQUEST
            ),
        )


class ReferralViewSet(viewsets.ReadOnlyModelViewSet):
    queryset = Referral.objects.select_related("referrer", "referred").all()
    serializer_class = ReferralSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        phone = self.request.query_params.get("phone")
        if phone:
            qs = qs.filter(
                Q(referrer__phone=phone) | Q(referred__phone=phone)
            )
        return qs


class NotificationViewSet(viewsets.ModelViewSet):
    serializer_class = NotificationSerializer

    def get_queryset(self):
        qs = LoyaltyNotification.objects.all()
        phone = self.request.query_params.get("phone")
        if phone:
            qs = qs.filter(customer__phone=phone)
        is_read = self.request.query_params.get("read")
        if is_read is not None:
            qs = qs.filter(is_read=is_read.lower() == "true")
        ntype = self.request.query_params.get("type")
        if ntype:
            qs = qs.filter(notification_type=ntype)
        return qs

    @action(detail=False, methods=["post"], url_path="mark-read")
    def mark_read(self, request):
        ser = NotificationMarkReadSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        phone = (
            request.data.get("phone")
            or request.headers.get("X-Customer-Phone")
        )
        qs = LoyaltyNotification.objects.filter(is_read=False)
        if phone:
            qs = qs.filter(customer__phone=phone)
        if ser.validated_data.get("mark_all"):
            count = qs.update(is_read=True)
        else:
            count = qs.filter(
                id__in=ser.validated_data["notification_ids"]
            ).update(is_read=True)
        return Response({"marked_read": count})


class LoyaltyTransactionViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = LoyaltyTransactionSerializer

    def get_queryset(self):
        qs = LoyaltyTransaction.objects.select_related("customer").all()
        phone = self.request.query_params.get("phone")
        if phone:
            qs = qs.filter(customer__phone=phone)
        ttype = self.request.query_params.get("type")
        if ttype:
            qs = qs.filter(transaction_type=ttype)
        return qs


class RewardRedemptionViewSet(viewsets.ReadOnlyModelViewSet):
    serializer_class = RewardRedemptionSerializer

    def get_queryset(self):
        qs = RewardRedemption.objects.select_related(
            "customer", "reward"
        ).all()
        phone = self.request.query_params.get("phone")
        if phone:
            qs = qs.filter(customer__phone=phone)
        return qs


# ══════════════════════════════════════════════════════════════════════════════
#  LOYALTY — FUNCTION-BASED API VIEWS
# ══════════════════════════════════════════════════════════════════════════════


@api_view(["POST"])
def process_order_loyalty_view(request):
    ser = ProcessOrderLoyaltySerializer(data=request.data)
    ser.is_valid(raise_exception=True)
    return Response(process_order_loyalty(**ser.validated_data))


@api_view(["GET"])
def loyalty_dashboard_view(request):
    return Response(
        LoyaltyDashboardSerializer(get_loyalty_dashboard()).data
    )


@api_view(["POST"])
def birthday_check_view(request):
    return Response({"birthday_granted": run_birthday_check_all()})


@api_view(["POST"])
def seed_levels_view(request):
    return Response({"message": seed_membership_levels()})


# ══════════════════════════════════════════════════════════════════════════════
#  LOYALTY — PAGE VIEWS (HTML)
# ══════════════════════════════════════════════════════════════════════════════


@staff_member_required
def loyalty_dashboard_page(request: HttpRequest):
    return render(
        request,
        "loyalty/dashboard.html",
        {"stats": get_loyalty_dashboard()},
    )


@staff_member_required
def loyalty_customers_page(request: HttpRequest):
    return render(request, "loyalty/customers.html")


@staff_member_required
def loyalty_customer_detail_page(request: HttpRequest, pk: int):
    return render(request, "loyalty/customer_detail.html")


@staff_member_required
def loyalty_coupons_page(request: HttpRequest):
    return render(request, "loyalty/coupons.html")


@staff_member_required
def loyalty_rewards_page(request: HttpRequest):
    return render(request, "loyalty/rewards.html")


@staff_member_required
def loyalty_notifications_page(request: HttpRequest):
    return render(
        request,
        "loyalty/notifications.html",
        {
            "unread_notifications": LoyaltyNotification.objects.filter(
                is_read=False
            ).count()
        },
    )


@staff_member_required
def loyalty_register_page(request: HttpRequest):
    return render(request, "loyalty/register.html")


# ══════════════════════════════════════════════════════════════════════════════
#  AUTHENTICATION VIEWS
# ══════════════════════════════════════════════════════════════════════════════


class LoginView(TokenObtainPairView):
    serializer_class = CustomTokenObtainSerializer


class RefreshView(TokenRefreshView):
    pass


class RegisterView(generics.CreateAPIView):
    serializer_class = RegisterSerializer
    permission_classes = [permissions.AllowAny]

    def create(self, request, *args, **kwargs):
        ser = self.get_serializer(data=request.data)
        ser.is_valid(raise_exception=True)
        result = register_user(ser.validated_data)

        # ← اگر کاربر تأیید نشده، توکن نده
        if not result["user"].is_approved:
            return api_success(
                data={"pending": True},
                message="ثبت‌نام شما با موفقیت انجام شد. پس از تأیید مدیر می‌توانید وارد شوید.",
                status_code=201,
            )

        return api_success(
            data={
                "user": UserDetailSerializer(result["user"]).data,
                "tokens": result["tokens"],
            },
            message=result["message"],
            status_code=201,
        )


class LogoutView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        try:
            refresh_token = request.data.get("refresh")
            if not refresh_token:
                return api_error("refresh token الزامی است.")
            from rest_framework_simplejwt.tokens import RefreshToken

            RefreshToken(refresh_token).blacklist()
            return api_success(message="خروج موفقیت‌آمیز بود.")
        except Exception:
            return api_error("توکن نامعتبر است.")


class CurrentUserView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        return api_success(data=UserDetailSerializer(request.user).data)

    def patch(self, request):
        ser = ProfileSerializer(request.user, data=request.data, partial=True)
        ser.is_valid(raise_exception=True)
        ser.save()
        return api_success(data=ser.data, message="پروفایل بروزرسانی شد.")


class ChangePasswordView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        ser = ChangePasswordSerializer(
            data=request.data, context={"request": request}
        )
        ser.is_valid(raise_exception=True)
        result = change_password(
            request.user,
            ser.validated_data["old_password"],
            ser.validated_data["new_password"],
        )
        if result["success"]:
            return api_success(message=result["message"])
        return api_error(result["error"])


class ResetPasswordView(APIView):
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        ser = ResetPasswordSerializer(data=request.data)
        ser.is_valid(raise_exception=True)
        result = reset_password(
            ser.validated_data["phone_number"],
            ser.validated_data["new_password"],
        )
        if result["success"]:
            return api_success(message=result["message"])
        return api_error(result["error"])


class UserListView(generics.ListAPIView):
    serializer_class = UserListSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        qs = AuthUser.objects.filter(
            restaurant=self.request.user.restaurant
        )
        role = self.request.query_params.get("role")
        if role:
            qs = qs.filter(role=role)
        search = self.request.query_params.get("search")
        if search:
            qs = qs.filter(
                Q(username__icontains=search)
                | Q(phone_number__icontains=search)
            )
        return qs


class UserDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = UserDetailSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        return AuthUser.objects.filter(
            restaurant=self.request.user.restaurant
        )
# ═══ Product Category Lookup ═══
def product_category_lookup(request):
    """جستجوی دسته‌بندی ذخیره‌شده بر اساس نام محصول"""
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({})

    item = (
        PurchaseInvoiceItem.objects
        .filter(item_name__iexact=q, category__isnull=False)
        .select_related('category')
        .order_by('-invoice__created_at')
        .first()
    )

    if item and item.category:
        return JsonResponse({
            'found': True,
            'category_id': item.category.id,
            'category_name': item.category.name,
        })

    return JsonResponse({'found': False})




# ═══════════════════════════════════════════
#  بستن روز — خلاصه (آپدیت شده)
# ═══════════════════════════════════════════

@csrf_exempt
@login_required
def pos_close_summary(request):
    """خلاصه قبل از بستن روز"""
    today = timezone.localdate()
    orders = Order.objects.filter(created_at__date=today)

    total_sales = orders.aggregate(s=Sum('total_price'))['s'] or 0
    order_count = orders.count()
    delivered = orders.filter(status='delivered').count()
    pending = orders.exclude(status='delivered').count()

    # سفارشات باز
    pending_orders = []
    for o in orders.exclude(status='delivered'):
        pending_orders.append({
            'id': o.id,
            'customer': o.customer_name or 'بدون نام',
            'total': o.total_price,
            'items': [
                {'name': oi.food.name if oi.food else '?', 'qty': oi.quantity}
                for oi in o.items.all()
            ]
        })

    # محصولات آشپزخانه (برای dropdown ضایعات)
    kitchen_items = []
    for kp in KitchenProduct.objects.filter(is_active=True):
        inv = kp.get_inventory()
        kitchen_items.append({
            'id': kp.id,
            'name': kp.name,
            'stock': inv.quantity,
            'category': kp.category,
        })

    # ضایعات امروز
    waste_logs = WasteLog.objects.filter(created_at__date=today)
    waste_count = waste_logs.aggregate(s=Sum('quantity'))['s'] or 0
    waste_value = 0
    for wl in waste_logs:
        kp = KitchenProduct.objects.filter(id=wl.kitchen_product_id).first()
        if kp:
            waste_value += (kp.selling_price or 0) * wl.quantity

    # تخفیف‌ها
    discount_total = 0
    for o in orders:
        if hasattr(o, 'discount_amount') and o.discount_amount:
            discount_total += o.discount_amount

    # آیتم‌های فروخته‌شده
    items_detail = []
    item_stats = {}
    for oi in OrderItem.objects.filter(order__created_at__date=today):
        name = oi.food.name if oi.food else '?'
        if name not in item_stats:
            item_stats[name] = {'qty': 0, 'revenue': 0}
        item_stats[name]['qty'] += oi.quantity
        item_stats[name]['revenue'] += oi.price * oi.quantity
    for name, stats in item_stats.items():
        items_detail.append({
            'name': name,
            'qty': stats['qty'],
            'revenue': stats['revenue'],
        })

    # پرفروش‌ترین‌ها
    top_items = sorted(items_detail, key=lambda x: x['qty'], reverse=True)[:5]

    # هزینه و سود
    total_cost = 0
    for oi in OrderItem.objects.filter(order__created_at__date=today):
        kp = KitchenProduct.objects.filter(recipe__food=oi.food).first()
        if kp and kp.recipe:
            # هزینه مواد اولیه × تعداد
            total_cost += 0  # اگه cost_price اضافه بشه اینجا محاسبه میشه
    total_profit = total_sales - total_cost - waste_value - discount_total

    # بررسی قبلاً بسته شده؟
    existing_report = DayCloseReport.objects.filter(date=today).first()

    return JsonResponse({
        'success': True,
        'total_sales': total_sales,
        'total_cost': total_cost,
        'total_profit': total_profit,
        'order_count': order_count,
        'delivered_count': delivered,
        'pending_count': pending,
        'pending_orders': pending_orders,
        'kitchen_items': kitchen_items,
        'waste_count': waste_count,
        'waste_value': waste_value,
        'discount_total': discount_total,
        'items_detail': items_detail,
        'top_items': top_items,
        'already_closed': existing_report is not None,
        'report_id': existing_report.id if existing_report else None,
    })


# ═══════════════════════════════════════════
#  بستن روز — ثبت ضایعات (بدون تغییر)
# ═══════════════════════════════════════════
@csrf_exempt
@login_required
def pos_register_waste(request):
    """ثبت ضایعات"""
    try:
        data = json_module.loads(request.body)
        items = data.get('items', [])
        if not items:
            return JsonResponse({'success': False, 'error': 'آیتمی ارسال نشد'})

        registered = []
        for item in items:
            kp_id = item.get('kitchen_product_id')
            qty = item.get('quantity', 0)
            note = item.get('note', '')

            # ★ تغییر: continue → خطا برگردون
            if qty <= 0:
                return JsonResponse({
                    'success': False,
                    'error': f'تعداد باید بیشتر از صفر باشد (مقدار: {qty})'
                })

            try:
                kp = KitchenProduct.objects.get(id=kp_id)
            except KitchenProduct.DoesNotExist:
                return JsonResponse({
                    'success': False,
                    'error': f'محصول آشپزخانه با شناسه {kp_id} پیدا نشد'
                })

            inv = kp.get_inventory()
            actual_qty = min(qty, inv.quantity)

            if actual_qty > 0:
                inv.quantity -= actual_qty
                inv.save(update_fields=['quantity', 'updated_at'])

                WasteLog.objects.create(
                    kitchen_product_id=kp.id,
                    quantity=actual_qty,
                    reason=note,
                )
                registered.append(f'{kp.name}×{actual_qty}')

        total = sum(1 for _ in registered)
        return JsonResponse({
            'success': True,
            'msg': f'ضایعات ثبت شد: {", ".join(registered)} (کل: {total})',
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
    
# ═══════════════════════════════════════════
#  بستن روز — تحویل همه (بدون تغییر)
# ═══════════════════════════════════════════

@csrf_exempt
@login_required
def pos_close_all_pending(request):
    """تحویل همه سفارشات باز"""
    today = timezone.localdate()
    pending = Order.objects.filter(
        created_at__date=today
    ).exclude(status='delivered')
    count = pending.count()
    pending.update(status='delivered')

    return JsonResponse({
        'success': True,
        'msg': f'{count} سفارش تحویل شد',
    })


# ═══════════════════════════════════════════
#  بستن روز — نهایی (آپدیت شده: ذخیره گزارش + لاگ)
# ═══════════════════════════════════════════

@csrf_exempt
@login_required
def pos_close_day(request):
    """بستن روز — تحویل همه + ذخیره گزارش نهایی + لاگ"""
    today = timezone.localdate()
    user = request.user if request.user.is_authenticated else None

    # ── ۱. تحویل همه سفارشات باز ──
    orders = Order.objects.filter(created_at__date=today)
    pending = orders.exclude(status='delivered')
    pending_count = pending.count()
    pending.update(status='delivered')

    # ── ۲. جمع‌آوری آمار ──
    total_sales = orders.aggregate(s=Sum('total_price'))['s'] or 0
    order_count = orders.count()
    delivered_count = orders.filter(status='delivered').count()

    waste_logs = WasteLog.objects.filter(created_at__date=today)
    waste_count = waste_logs.aggregate(s=Sum('quantity'))['s'] or 0
    waste_value = 0
    for wl in waste_logs:
        kp = KitchenProduct.objects.filter(id=wl.kitchen_product_id).first()
        if kp:
            waste_value += (kp.selling_price or 0) * wl.quantity

    discount_total = 0
    for o in orders:
        if hasattr(o, 'discount_amount') and o.discount_amount:
            discount_total += o.discount_amount

    items_detail = []
    item_stats = {}
    for oi in OrderItem.objects.filter(order__created_at__date=today):
        name = oi.food.name if oi.food else '?'
        if name not in item_stats:
            item_stats[name] = {'qty': 0, 'revenue': 0}
        item_stats[name]['qty'] += oi.quantity
        item_stats[name]['revenue'] += oi.price * oi.quantity
    for name, stats in item_stats.items():
        items_detail.append({'name': name, 'qty': stats['qty'], 'revenue': stats['revenue']})

    top_items = sorted(items_detail, key=lambda x: x['qty'], reverse=True)[:5]

    total_cost = 0
    total_profit = total_sales - total_cost - waste_value - discount_total

    inventory_snapshot = {}
    for kp in KitchenProduct.objects.filter(is_active=True):
        inv = kp.get_inventory()
        inventory_snapshot[kp.name] = {
            'product_id': kp.id,
            'stock': inv.quantity,
            'price': kp.selling_price or 0,
        }

    # ★FIX: objects.create بجای update_or_create — هر بستن report جدید می‌سازه
    report = DayCloseReport.objects.create(
        date=today,
        total_sales=total_sales,
        total_cost=total_cost,
        total_profit=total_profit,
        order_count=order_count,
        delivered_count=delivered_count,
        waste_count=waste_count,
        waste_value=waste_value,
        discount_total=discount_total,
        inventory_snapshot=inventory_snapshot,
        items_detail=items_detail,
        top_items=top_items,
        closed_by=user,
    )

    DayCloseLog.objects.create(
        date=today,
        action='close',
        user=user,
        details={
            'report_id': report.id,
            'total_sales': total_sales,
            'order_count': order_count,
            'waste_count': waste_count,
            'pending_delivered': pending_count,
        },
    )

    return JsonResponse({
        'success': True,
        'report_id': report.id,
        'msg': f'روز بسته شد — {order_count} سفارش / '
               f'{total_sales:,} تومان فروش / '
               f'{total_profit:,} تومان سود / '
               f'{delivered_count} تحویل / '
               f'{waste_count} ضایعات',
    })

# ═══════════════════════════════════════════
#  تاریخچه بستن روز — API جدید
# ═══════════════════════════════════════════

@login_required
def pos_close_history(request):
    """لیست گزارش‌های بستن روز"""
    limit = int(request.GET.get('limit', 30))
    reports = DayCloseReport.objects.all()[:limit]
    data = []
    for r in reports:
        data.append({
            'id': r.id,
            'date': str(r.date),
            'total_sales': r.total_sales,
            'total_cost': r.total_cost,
            'total_profit': r.total_profit,
            'order_count': r.order_count,
            'delivered_count': r.delivered_count,
            'waste_count': r.waste_count,
            'waste_value': r.waste_value,
            'discount_total': r.discount_total,
            'closed_by': r.closed_by.username if r.closed_by else '?',
            'closed_at': r.closed_at.strftime('%Y-%m-%d %H:%M'),
        })
    return JsonResponse({'success': True, 'reports': data})


@login_required
def pos_close_report_detail(request, report_id):
    """جزئیات یک گزارش بستن روز"""
    try:
        r = DayCloseReport.objects.get(id=report_id)
    except DayCloseReport.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'گزارش پیدا نشد'})

    return JsonResponse({
        'success': True,
        'report': {
            'id': r.id,
            'date': str(r.date),
            'total_sales': r.total_sales,
            'total_cost': r.total_cost,
            'total_profit': r.total_profit,
            'order_count': r.order_count,
            'delivered_count': r.delivered_count,
            'waste_count': r.waste_count,
            'waste_value': r.waste_value,
            'discount_total': r.discount_total,
            'inventory_snapshot': r.inventory_snapshot,
            'items_detail': r.items_detail,
            'top_items': r.top_items,
            'closed_by': r.closed_by.username if r.closed_by else '?',
            'closed_at': r.closed_at.strftime('%Y-%m-%d %H:%M'),
        }
    })


@login_required
def pos_close_logs(request):
    """لاگ‌های بستن/بازکردن روز"""
    limit = int(request.GET.get('limit', 50))
    logs = DayCloseLog.objects.select_related('user').all()[:limit]
    data = []
    for log in logs:
        data.append({
            'id': log.id,
            'date': str(log.date),
            'action': log.action,
            'action_display': log.get_action_display(),
            'user': log.user.username if log.user else '?',
            'details': log.details,
            'created_at': log.created_at.strftime('%Y-%m-%d %H:%M:%S'),
        })
    return JsonResponse({'success': True, 'logs': data})

# ═══════════════════════════════════════════════════════
#  Kitchen Waste CRUD — /api/kitchen/waste/
# ═══════════════════════════════════════════════════════
# ★ اضافه کن بالای کلاس — لیست دلایل مجاز
VALID_WASTE_REASONS = [
    'expired', 'damaged', 'overcooked',
    'quality_issue', 'returned', 'other',
]


class KitchenWasteListCreate(APIView):
    """GET: لیست ضایعات / POST: ثبت ضایعات جدید"""
    permission_classes = [IsAuthenticated]

    def get(self, request):
        logs = WasteLog.objects.select_related('kitchen_product').order_by('-created_at')
        data = []
        for w in logs:
            kp = w.kitchen_product
            data.append({
                'id': w.id,
                'kitchen_product': w.kitchen_product_id,
                'kitchen_product_name': kp.name if kp else '?',
                'quantity': w.quantity,
                'reason': getattr(w, 'reason', '') or '',
                'created_at': w.created_at.isoformat() if hasattr(w, 'created_at') else '',
            })
        return Response(data)

    def post(self, request):
        d = request.data
        kp_id = d.get('kitchen_product')
        qty = d.get('quantity', 0)
        reason = d.get('reason')  # ★ تغییر: default حذف شد

        if not kp_id:
            return Response({'error': 'محصول مشخص نشده'}, status=400)
        if not isinstance(qty, (int, float)) or qty <= 0:
            return Response({'error': 'تعداد باید بزرگتر از صفر باشد'}, status=400)

        # ★ اضافه شد: validation دلیل
        if not reason:
            return Response({'error': 'دلیل ضایعات الزامی است'}, status=400)
        if reason not in VALID_WASTE_REASONS:
            return Response({'error': f'دلیل نامعتبر: {reason}'}, status=400)

        try:
            kp = KitchenProduct.objects.get(id=kp_id)
        except KitchenProduct.DoesNotExist:
            return Response({'error': 'محصول یافت نشد'}, status=404)

        inv = kp.get_inventory()
        actual_qty = min(qty, inv.quantity)
        if actual_qty <= 0:
            return Response({'error': 'موجودی صفر است'}, status=400)

        waste = WasteLog.objects.create(
            kitchen_product=kp,
            quantity=actual_qty,
            reason=reason,
        )

        inv.quantity -= actual_qty
        if inv.quantity < 0:
            inv.quantity = 0
        inv.save(update_fields=['quantity', 'updated_at'])

        return Response({
            'id': waste.id,
            'kitchen_product': kp.id,
            'quantity': actual_qty,
            'reason': reason,
        }, status=201)
    
class KitchenWasteDetail(APIView):
    """DELETE: حذف ضایعات"""
    permission_classes = [IsAuthenticated]

    def delete(self, request, pk):
        try:
            w = WasteLog.objects.get(id=pk)
        except WasteLog.DoesNotExist:
            return Response({'error': 'یافت نشد'}, status=404)

        inv = w.kitchen_product.get_inventory()
        inv.quantity += w.quantity
        inv.save(update_fields=['quantity', 'updated_at'])
        w.delete()
        return Response({'success': True})
# ═══════════════════════════════════════════════════════
#  منوی عمومی — API برای فرانت‌اند مشتری
# ═══════════════════════════════════════════════════════
@api_view(['GET'])
@permission_classes([AllowAny])
def public_food_list(request):
    """لیست غذاها برای منوی عمومی"""
    foods = Food.objects.all().select_related('category')

    category_id = request.query_params.get('category')
    if category_id:
        foods = foods.filter(category_id=category_id)

    data = []
    for food in foods:
        data.append({
            'id': food.id,
            'name': food.name,
            'description': '',
            'price': int(food.final_price),
            'final_price': int(food.final_price),
            'image': food.image.url if food.image else '',
            'category': food.category_id,
            'category_name': food.category.name if food.category else '',
        })

    return Response(data)


@api_view(['GET'])
@permission_classes([AllowAny])
def public_category_list(request):
    """لیست دسته‌بندی‌ها برای منوی عمومی"""
    categories = Category.objects.filter(is_active=True).order_by('order', 'name')

    data = []
    for cat in categories:
        data.append({
            'id': cat.id,
            'name': cat.name,
            'image': cat.image.url if cat.image else '',
        })

    return Response(data)

@api_view(["POST"])
@staff_member_required
def order_send_to_kitchen(request, pk):
    """
    سفارش رو به آشپزخانه میفرسته:
    1. رسپی هر غذا رو پیدا میکنه
    2. مواد اولیه (مستقیم + نیم‌آماده) رو بررسی و کم میکنه
    3. InventoryMovement و InventoryUsageLog ثبت میکنه
    4. وضعیت سفارش رو تغییر میده
    """

    # ── پیدا کردن سفارش ──
    try:
        order = Order.objects.prefetch_related(
            'items__food__recipe__ingredients__raw_material',
            'items__food__recipe__semi_finished_items__semi_finished__ingredients__raw_material',
        ).get(pk=pk)
    except Order.DoesNotExist:
        return JsonResponse({"error": "سفارش یافت نشد."}, status=404)

    # ── بررسی وضعیت ──
    if order.status not in ('pending', 'confirmed'):
        return JsonResponse(
            {"error": "فقط سفارشات «در انتظار» یا «تأیید شده» قابل ارسال هستند."},
            status=400,
        )

    errors = []
    materials_used = []

    with transaction.atomic():

        # ═══════════════════════════════════
        #  مرحله ۱: بررسی موجودی (بدون کسر)
        # ═══════════════════════════════════
        for item in order.items.all():
            food = item.food
            if not food:
                continue

            qty = item.quantity

            # آیا رسپی داره؟
            try:
                recipe = food.recipe
            except Recipe.DoesNotExist:
                errors.append(f"«{food.name}» رسپی ندارد")
                continue

            # ── بررسی مواد اولیه مستقیم ──
            for ri in recipe.ingredients.all():
                needed = Decimal(str(ri.effective_quantity)) * qty
                available = ri.raw_material.quantity
                if available < needed:
                    errors.append(
                        f"«{food.name}»: {ri.raw_material.name} "
                        f"کم است (نیاز: {needed} {ri.raw_material.get_unit_display()}، "
                        f"موجود: {available} {ri.raw_material.get_unit_display()})"
                    )

            # ── بررسی مواد نیم‌آماده ──
            for rsf in recipe.semi_finished_items.all():
                sf = rsf.semi_finished
                needed_sf = Decimal(str(rsf.quantity)) * qty

                # موجودی نیم‌آماده کافیه؟
                if sf.current_stock < needed_sf:
                    errors.append(
                        f"«{food.name}»: نیم‌آماده «{sf.name}» "
                        f"کم است (نیاز: {needed_sf}، موجود: {sf.current_stock})"
                    )
                    continue

                # مواد اولیه نیم‌آماده کافیه؟
                for sfi in sf.ingredients.all():
                    needed_raw = sfi.quantity * needed_sf
                    if sfi.raw_material.quantity < needed_raw:
                        errors.append(
                            f"«{food.name}» ← «{sf.name}»: "
                            f"{sfi.raw_material.name} کم است "
                            f"(نیاز: {needed_raw} {sfi.raw_material.get_unit_display()}، "
                            f"موجود: {sfi.raw_material.quantity} {sfi.raw_material.get_unit_display()})"
                        )

        # اگه خطا بود → برگرد (هیچی کم نشده)
        if errors:
            return JsonResponse({"error": errors}, status=400)

        # ═══════════════════════════════════
        #  مرحله ۲: کسر مواد اولیه
        # ═══════════════════════════════════
        for item in order.items.all():
            food = item.food
            if not food:
                continue

            qty = item.quantity
            recipe = food.recipe

            # ── کسر مواد اولیه مستقیم ──
            for ri in recipe.ingredients.all():
                needed = Decimal(str(ri.effective_quantity)) * qty
                rm = ri.raw_material
                prev_stock = rm.quantity
                rm.quantity -= needed
                rm.save(update_fields=['quantity'])

                InventoryMovement.objects.create(
                    raw_material=rm,
                    movement_type='order_usage',
                    quantity=needed,
                    previous_stock=prev_stock,
                    new_stock=rm.quantity,
                    reference_type='order',
                    reference_id=order.id,
                    notes=f'سفارش #{order.id} — {food.name} ×{qty}',
                    created_by=request.user,
                )

                InventoryUsageLog.objects.create(
                    raw_material=rm,
                    usage_type='order',
                    quantity_used=needed,
                    reference=f'سفارش #{order.id}',
                    note=f'{food.name} ×{qty}',
                )

                materials_used.append({
                    'name': rm.name,
                    'quantity': float(needed),
                    'unit': rm.get_unit_display(),
                    'type': 'direct',
                })

            # ── کسر مواد نیم‌آماده و مواد اولیه‌شون ──
            for rsf in recipe.semi_finished_items.all():
                sf = rsf.semi_finished
                needed_sf = Decimal(str(rsf.quantity)) * qty

                # کسر موجودی نیم‌آماده
                sf.current_stock -= needed_sf
                sf.save(update_fields=['current_stock'])

                # کسر مواد اولیه نیم‌آماده
                for sfi in sf.ingredients.all():
                    needed_raw = sfi.quantity * needed_sf
                    rm = sfi.raw_material
                    prev_stock = rm.quantity
                    rm.quantity -= needed_raw
                    rm.save(update_fields=['quantity'])

                    InventoryMovement.objects.create(
                        raw_material=rm,
                        movement_type='order_usage',
                        quantity=needed_raw,
                        previous_stock=prev_stock,
                        new_stock=rm.quantity,
                        reference_type='order',
                        reference_id=order.id,
                        notes=f'سفارش #{order.id} — {sf.name} ← {food.name} ×{qty}',
                        created_by=request.user,
                    )

                    InventoryUsageLog.objects.create(
                        raw_material=rm,
                        usage_type='order',
                        quantity_used=needed_raw,
                        reference=f'سفارش #{order.id}',
                        note=f'{sf.name} ← {food.name} ×{qty}',
                    )

                    materials_used.append({
                        'name': rm.name,
                        'quantity': float(needed_raw),
                        'unit': rm.get_unit_display(),
                        'type': f'semi:{sf.name}',
                    })

        # ═══════════════════════════════════
        #  مرحله ۳: تغییر وضعیت سفارش
        # ═══════════════════════════════════
        order.status = 'preparing'
        order.save(update_fields=['status'])

    return JsonResponse({
        "success": True,
        "msg": f"سفارش #{order.id} به آشپزخانه ارسال شد.",
        "materials_used": materials_used,
    })



@api_view(["GET"])
@staff_member_required
def kitchen_orders_api(request):
    """سفارشاتی که باید آشپزخانه آماده کنه"""
    status_filter = request.GET.get("status", "preparing")

    orders = Order.objects.prefetch_related(
        'items__food'
    ).filter(
        status=status_filter
    ).order_by('created_at')[:50]

    data = []
    for order in orders:
        items = []
        for item in order.items.all():
            items.append({
                "food_name": item.food.name if item.food else "—",
                "quantity": item.quantity,
            })

        data.append({
            "id": order.id,
            "status": order.status,
            "customer_name": order.customer_name or "—",
            "items": items,
            "total_price": int(order.total_price),
            "created_at": order.created_at.strftime("%H:%M"),
        })

    return JsonResponse({"orders": data}, safe=False)



CARD_READER_URL = f"http://{getattr(settings, 'CARD_READER_IP', '127.0.0.1')}:{getattr(settings, 'CARD_READER_PORT', 8080)}"

@csrf_exempt
@require_POST
def send_to_card_reader(request):
    """ارسال مبلغ به کارتخوان"""
    try:
        data = json.loads(request.body)
        amount = data.get('amount')
        order_id = data.get('order_id')

        if not amount or amount <= 0:
            return JsonResponse({'success': False, 'error': 'مبلغ نامعتبر'})

        resp = requests.post(
            f"{CARD_READER_URL}/api/payment",
            json={
                'amount': int(amount),
                'rrn': str(order_id),
                'description': f'سفارش #{order_id}',
            },
            timeout=120
        )
        result = resp.json()

        if result.get('status') == 'approved' or result.get('success'):
            card_num = result.get('card_number', '')
            return JsonResponse({
                'success': True,
                'trace_number': result.get('trace_number', ''),
                'ref_number': result.get('ref_number', ''),
                'card_last4': card_num[-4:] if card_num else '',
                'message': 'پرداخت موفق'
            })
        else:
            return JsonResponse({
                'success': False,
                'error': result.get('message', 'پرداخت ناموفق')
            })

    except requests.Timeout:
        return JsonResponse({'success': False, 'error': 'زمان انتظار تمام شد'})
    except requests.ConnectionError:
        return JsonResponse({'success': False, 'error': f'کارتخوان ({CARD_READER_URL}) در دسترس نیست'})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@csrf_exempt
@require_POST
def cancel_card_payment(request):
    """لغو پرداخت"""
    try:
        resp = requests.post(f"{CARD_READER_URL}/api/payment/cancel", timeout=10)
        return JsonResponse({'success': True})
    except:
        return JsonResponse({'success': False})
    
# ══════════════════════════════════════════════════════════════
#  User Management — مدیریت کاربران و نقش‌ها
# ══════════════════════════════════════════════════════════════

from .roles import (
    ROLE_CHOICES, get_user_role, get_user_permissions,
    get_role_display, has_permission, require_permission,
)


@staff_member_required
def user_management_page(request: HttpRequest):
    """صفحه مدیریت کاربران — با مجوزها."""
    roles = [
        {
            "value": "owner",
            "label": "مالک",
            "permissions": [
                "foods.view", "foods.edit", "foods.create", "foods.delete", "foods.categories",
                "inventory.view", "inventory.edit", "inventory.create", "inventory.delete",
                "inventory.raw_materials", "inventory.ready_materials", "inventory.semi_finished",
                "inventory.usages_log", "inventory.invoice", "inventory.end_of_invoice",
                "orders.view", "orders.edit", "orders.create", "orders.delete",
                "pos.view", "pos.use", "pos.close", "pos.report",
                "kitchen.view", "kitchen.manage",
                "loyalty.view", "loyalty.edit", "loyalty.customers", "loyalty.coupons", "loyalty.rewards",
                "users.view", "users.edit", "users.create", "users.delete",
            ],
        },
        {
            "value": "manager",
            "label": "مدیر",
            "permissions": [
                "foods.view", "foods.edit", "foods.categories",
                "inventory.view", "inventory.edit", "inventory.raw_materials",
                "inventory.ready_materials", "inventory.usages_log", "inventory.invoice",
                "orders.view", "orders.edit", "orders.create",
                "pos.view", "pos.use", "pos.close", "pos.report",
                "kitchen.view", "kitchen.manage",
                "loyalty.view", "loyalty.customers",
            ],
        },
        {
            "value": "cashier",
            "label": "صندوقدار",
            "permissions": [
                "foods.view",
                "orders.view", "orders.create",
                "pos.view", "pos.use", "pos.close",
                "loyalty.view", "loyalty.customers",
            ],
        },
        {
            "value": "kitchen",
            "label": "آشپز",
            "permissions": [
                "foods.view",
                "kitchen.view", "kitchen.manage",
            ],
        },
        {
            "value": "waiter",
            "label": "پیشخدمت",
            "permissions": [
                "foods.view",
                "orders.view", "orders.create",
            ],
        },
    ]

    return render(request, "restaurant/user_management.html", {
        "roles_json": json_module.dumps(roles, ensure_ascii=False),
        "current_user_id": request.user.id,
    })

# ═══ API: لیست کاربران ═══
@api_view(["GET"])
@permission_classes([IsAuthenticated])
def user_management_api(request):
    if not has_permission(request.user, "users.view"):
        return Response({"error": "دسترسی ندارید."}, status=403)

    # هم کاربران رستوران فعلی، هم کاربران بدون رستوران (در انتظار تأیید)
    from django.db.models import Q
    users = AuthUser.objects.filter(
        Q(restaurant=request.user.restaurant) | Q(restaurant__isnull=True)
    ).order_by("-date_joined")

    data = []
    for u in users:
        data.append({
            "id": u.id,
            "username": u.username,
            "phone_number": getattr(u, "phone_number", "") or "",
            "first_name": getattr(u, "first_name", "") or "",
            "last_name": getattr(u, "last_name", "") or "",
            "is_approved": getattr(u, "is_approved", True),
            "role": getattr(u, "role", "cashier") or "cashier",
            "role_display": get_role_display(u),
            "is_active": u.is_active,
            "is_staff": u.is_staff,
            "date_joined": u.date_joined.strftime("%Y/%m/%d %H:%M") if u.date_joined else "—",
            "last_login": u.last_login.strftime("%Y/%m/%d %H:%M") if u.last_login else "هرگز",
            "permissions": get_user_permissions(u),
        })

    return Response({"success": True, "users": data})
# ═══ API: ایجاد کاربر ═══
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def create_user_api(request):
    if not has_permission(request.user, "users.create"):
        return Response({"success": False, "error": "دسترسی ندارید."}, status=403)

    username = request.data.get("username", "").strip()
    password = request.data.get("password", "")
    phone = request.data.get("phone_number", "").strip()
    role = request.data.get("role", "cashier")

    if not username:
        return Response({"success": False, "error": "نام کاربری الزامی است."}, status=400)
    if not password or len(password) < 4:
        return Response({"success": False, "error": "رمز باید حداقل 4 کاراکتر باشد."}, status=400)
    if AuthUser.objects.filter(username=username).exists():
        return Response({"success": False, "error": "این نام کاربری قبلاً ثبت شده."}, status=400)

    try:
        user = AuthUser.objects.create_user(username=username, password=password)
        if hasattr(user, "role"):
            user.role = role
        if hasattr(user, "phone_number"):
            user.phone_number = phone
        if hasattr(user, "restaurant"):
            user.restaurant = request.user.restaurant
        user.is_staff = True  
        user.is_superuser = False  # FIXED                  # ← همه کاربران مدیریتی staff باشن
        user.is_approved = False
        user.save()

        return Response({
            "success": True,
            "user_id": user.id,                 # ← اضافه شد
            "username": user.username,
            "msg": f"کاربر «{username}» ایجاد شد. پس از تأیید می‌تواند وارد شود.",
        })
    except Exception as e:
        return Response({"success": False, "error": str(e)}, status=500)

# ═══ API: تغییر نقش ═══
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def user_update_role(request):
    if not has_permission(request.user, "users.edit"):
        return Response({"error": "دسترسی ندارید."}, status=403)

    user_id = request.data.get("user_id")
    new_role = request.data.get("role")

    if not user_id or not new_role:
        return Response({"error": "شناسه کاربر و نقش الزامی است."}, status=400)

    valid_roles = [r[0] for r in ROLE_CHOICES]
    if new_role not in valid_roles:
        return Response({"error": "نقش نامعتبر."}, status=400)

    if int(user_id) == request.user.id:
        return Response({"error": "نمی‌توانید نقش خودتان را تغییر دهید."}, status=400)

    from .roles import ROLE_OWNER, ROLE_MANAGER
    if new_role == ROLE_OWNER and get_user_role(request.user) != ROLE_OWNER:
        return Response({"error": "فقط مالک می‌تواند مالک تعیین کند."}, status=403)

    try:
        target = AuthUser.objects.get(id=user_id, restaurant=request.user.restaurant)
    except AuthUser.DoesNotExist:
        return Response({"error": "کاربر یافت نشد."}, status=404)

    old_role = getattr(target, "role", "")
    target.role = new_role
    target.is_staff = new_role in (ROLE_OWNER, ROLE_MANAGER)
    target.save(update_fields=["role", "is_staff"])

    role_names = dict(ROLE_CHOICES)
    return Response({
        "success": True,
        "msg": f"نقش «{target.username}» از «{role_names.get(old_role, 'نامشخص')}» به «{role_names.get(new_role, 'نامشخص')}» تغییر کرد.",
        "user": {
            "id": target.id,
            "username": target.username,
            "role": new_role,
            "role_display": get_role_display(target),
        },
    })




# ═══ API: فعال/غیرفعال ═══
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def user_toggle_active(request):
    if not has_permission(request.user, "users.edit"):
        return Response({"error": "دسترسی ندارید."}, status=403)

    user_id = request.data.get("user_id")
    if not user_id:
        return Response({"error": "شناسه کاربر الزامی."}, status=400)

    if int(user_id) == request.user.id:
        return Response({"error": "نمی‌توانید خودتان را غیرفعال کنید."}, status=400)

    try:
        target = AuthUser.objects.get(id=user_id, restaurant=request.user.restaurant)
    except AuthUser.DoesNotExist:
        return Response({"error": "کاربر یافت نشد."}, status=404)

    target.is_active = not target.is_active
    target.save(update_fields=["is_active"])

    return Response({
        "success": True,
        "is_active": target.is_active,
        "msg": f"کاربر «{target.username}» {'فعال' if target.is_active else 'غیرفعال'} شد.",
    })


# ═══ API: تغییر رمز ═══
@api_view(["POST"])
@permission_classes([IsAuthenticated])
def admin_reset_password(request):
    if not has_permission(request.user, "users.edit"):
        return Response({"error": "دسترسی ندارید."}, status=403)

    user_id = request.data.get("user_id")
    new_password = request.data.get("new_password")

    if not user_id or not new_password:
        return Response({"error": "شناسه کاربر و رمز جدید الزامی است."}, status=400)

    if len(new_password) < 4:
        return Response({"error": "رمز عبور باید حداقل 4 کاراکتر باشد."}, status=400)

    if int(user_id) == request.user.id:
        return Response({"error": "برای تغییر رمز خودتان از بخش تغییر رمز استفاده کنید."}, status=400)

    try:
        target = AuthUser.objects.get(id=user_id, restaurant=request.user.restaurant)
    except AuthUser.DoesNotExist:
        return Response({"error": "کاربر یافت نشد."}, status=404)

    target.set_password(new_password)
    target.save(update_fields=["password"])

    return Response({
        "success": True,
        "msg": f"رمز «{target.username}» تغییر کرد.",
    })


@csrf_protect
@require_POST
def approve_user_api(request):
    """تأیید کاربر جدید و تعیین نقش."""

    # ↓↓↓ این ۳ خط رو اضافه کن ↓↓↓
    if not request.user.is_authenticated:
        return JsonResponse({"success": False, "error": "غیرمجاز"}, status=403)

    try:
        data = json.loads(request.body)
        user_id = data.get("user_id")
        role = data.get("role", "customer")

        if not user_id:
            return JsonResponse({"success": False, "error": "شناسه کاربر ارسال نشد."})

        user = User.objects.get(id=user_id)
        user.is_approved = True
        user.role = role
        user.is_staff = role in ("owner", "manager", "cashier", "kitchen")
        if hasattr(user, "restaurant"):
            user.restaurant = request.user.restaurant
        user.save()

        return JsonResponse({
            "success": True,
            "msg": f"کاربر «{user.username}» تأیید شد و نقش «{user.get_role_display()}» دریافت کرد.",
        })
    except User.DoesNotExist:
        return JsonResponse({"success": False, "error": "کاربر یافت نشد."})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})



@csrf_protect
@require_POST
def reject_user_api(request):
    """رد و حذف کاربر تأیید نشده."""
    try:
        data = json.loads(request.body)
        user_id = data.get("user_id")

        if not user_id:
            return JsonResponse({"success": False, "error": "شناسه کاربر ارسال نشد."})

        user = User.objects.get(id=user_id)
        if user.is_approved:
            return JsonResponse({"success": False, "error": "این کاربر قبلاً تأیید شده."})

        username = user.username
        user.delete()

        return JsonResponse({
            "success": True,
            "msg": f"درخواست «{username}» رد شد و حذف گردید.",
        })
    except User.DoesNotExist:
        return JsonResponse({"success": False, "error": "کاربر یافت نشد."})
    except Exception as e:
        return JsonResponse({"success": False, "error": str(e)})
    
from django.contrib.auth import login

class SetSessionView(APIView):
    """بعد از JWT login، Django session هم بساز تا صفحات HTML کار کنن"""
    permission_classes = [permissions.AllowAny]

    def post(self, request):
        access_token = request.data.get('access_token')
        user_id = request.data.get('user_id')

        if not user_id:
            return Response({'success': False}, status=400)

        try:
            user = User.objects.get(id=user_id)
            login(request, user, backend='django.contrib.auth.backends.ModelBackend')
            return Response({'success': True})
        except User.DoesNotExist:
            return Response({'success': False, 'error': 'کاربر یافت نشد'}, status=404)
        
#--------------------
@csrf_exempt
def user_delete(request):
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'error': 'غیرمجاز'}, status=403)

    if not (request.user.is_superuser or getattr(request.user, 'role', '') == 'owner'):
        return JsonResponse({'success': False, 'error': 'دسترسی غیرمجاز'}, status=403)

    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'درخواست نامعتبر'}, status=405)

    try:
        data = json_module.loads(request.body)
    except Exception:
        return JsonResponse({'success': False, 'error': 'داده نامعتبر'}, status=400)

    user_id = data.get('user_id')
    if not user_id:
        return JsonResponse({'success': False, 'error': 'شناسه کاربر الزامی است.'}, status=400)

    try:
        target = User.objects.get(id=user_id)
    except User.DoesNotExist:
        return JsonResponse({'success': False, 'error': 'کاربر یافت نشد.'}, status=400)

    if target.id == request.user.id:
        return JsonResponse({'success': False, 'error': 'نمی‌توانید خودتان را حذف کنید.'}, status=400)

    target.delete()
    return JsonResponse({'success': True})
